# app.py

# Standard library imports
import base64
import io
import json
import logging
import math
import os
import platform
import random
import re
import secrets
import time
import traceback
from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import wraps

# Third-party imports
import MySQLdb.cursors
import pdfkit
import qrcode
from flask import (Flask, abort, flash, g, jsonify, make_response, redirect,
                  render_template, request, send_file, session, url_for)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_login import current_user
from flask_mail import Mail, Message
from flask_mysqldb import MySQL
from flask_wtf.csrf import CSRFProtect
from itsdangerous import URLSafeTimedSerializer
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

# Local imports
from config import config
from forms import *
  
  

app = Flask(__name__)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Secret key for session
app.config.update(
    SECRET_KEY=os.urandom(24),
    WTF_CSRF_SECRET_KEY=os.urandom(24)  # for extra security
)

# Initialize CSRF protection AFTER setting secret key
csrf = CSRFProtect(app)

# MySQL configurations
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = 'Himal5221@'
app.config['MYSQL_DB'] = 'world_hotel_final_project_db'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'

# Initialize MySQL
mysql = MySQL(app)

# Initialize rate limiter and CSRF protection
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

csrf = CSRFProtect(app)

# Mail configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'acharyahimal122@gmail.com'
app.config['MAIL_PASSWORD'] = 'dsps yyig hbkq vleq'
app.config['MAIL_DEFAULT_SENDER'] = 'World Hotels acharyahimal122@gmail.com'

mail = Mail(app)

# Email sending function
def send_email(to, subject, template, **kwargs):
    """Generic email sending function"""
    try:
        msg = Message(
            subject,
            recipients=[to],
            html=render_template(template, **kwargs)  # Pass kwargs directly to template
        )
        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"Email sending error: {str(e)}")
        return False


# Email verification function
def send_verification_email(user_email, token):
    """Send email verification link"""
    verification_url = url_for('verify_email', token=token, _external=True)
    return send_email(
        to=user_email,
        subject='Verify Your Email - World Hotels',
        template='auth/verify_email.html',
        verification_url=verification_url
    )


# Password reset function
def send_password_reset_email(user_email, token):
    """Send password reset link"""
    try:
        reset_url = url_for('reset_password', token=token, _external=True)
        
        # Create and send the email
        msg = Message(
            'Password Reset Request - World Hotels',
            recipients=[user_email]
        )
        
        msg.html = render_template(
            'email/reset_password_email.html',
            reset_url=reset_url,
            email=user_email
        )
        
        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"Password reset email error: {str(e)}")
        return False

# Booking confirmation function
def send_booking_confirmation(user_email, booking_data):
    """Send booking confirmation email"""
    return send_email(
        to=user_email,
        subject='Booking Confirmation - World Hotels',
        template='email/booking_confirmation.html',
        booking=booking_data
    )

# Create upload directories if they don't exist
UPLOAD_FOLDERS = [
    'static/uploads',
    'static/images',
    'static/room_images',
    'static/hotel_images',
    'static/gallery_images'  # Added new directory for gallery
]

# Upload folders configuration
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'static/uploads')
app.config['HOTEL_IMAGES'] = os.path.join(app.root_path, 'static/hotel_images')
app.config['GALLERY_IMAGES'] = os.path.join(app.root_path, 'static/gallery_images')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# Allowed file extension function
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Create upload directories in not exists
for folder in UPLOAD_FOLDERS:
    os.makedirs(os.path.join(app.root_path, folder), exist_ok=True)


# Password strength validation function
def is_password_strong(password):
    """
    Validates if a password meets strength requirements:
    - At least 8 characters
    - At least one uppercase letter
    - At least one lowercase letter
    - At least one digit
    - At least one special character
    """
    regex = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&#])[A-Za-z\d@$!%*?&#]{8,}$'
    return re.match(regex, password)


# Utility Functions
def sanitize_input(text):
    """Basic input sanitization"""
    if text is None:
        return None
    return text.strip()


# Database Connection function
def get_db():
    """Safe database connection handling"""
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        return cursor
    except Exception as e:
        app.logger.error(f'Database connection error: {str(e)}')
        abort(500)

# Decorators
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.method == 'POST':
                # Generate a secure token
                token = secrets.token_urlsafe(32)
                
                try:
                    # Store form data temporarily in database
                    cursor = get_db()
                    cursor.execute('''
                        INSERT INTO temp_form_data 
                        (token, form_data, expiry) 
                        VALUES (%s, %s, DATE_ADD(NOW(), INTERVAL 10 MINUTE))
                    ''', (token, json.dumps(dict(request.form))))
                    mysql.connection.commit()
                    
                    # Store only the token in session
                    session['form_token'] = token
                    session['next_url'] = request.url
                    
                except Exception as e:
                    app.logger.error(f'Error storing form data: {str(e)}')
                finally:
                    if cursor:
                        cursor.close()
            
            flash('Please login first', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_role' not in session or session['user_role'] != 1:  # 1 is admin role_id
            flash('Admin access required', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function



# Error Handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    mysql.connection.rollback()
    return render_template('errors/500.html'), 500

@app.errorhandler(429)
def ratelimit_error(error):
    return render_template('errors/429.html'), 429

@app.errorhandler(403)
def csrf_error(error):
    return render_template('errors/403.html'), 403



# date filter for date format
@app.template_filter('date')
def date_filter(value, format='%Y-%m-%d'):
    if value is None:
        return ""
    if isinstance(value, str):
        value = datetime.strptime(value, '%Y-%m-%d')
    return value.strftime(format)


# time filter for time format
@app.template_filter('time')
def time_filter(value, format='%H:%M'):
    if value is None:
        return ""
    if isinstance(value, timedelta):
        # Convert timedelta to time string
        total_seconds = int(value.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, '%H:%M:%S')
        except ValueError:
            try:
                value = datetime.strptime(value, '%H:%M')
            except ValueError:
                return value
    return value.strftime(format)

# datetime filter for datetime format
@app.template_filter('datetime')
def datetime_filter(value, format='%Y-%m-%d %H:%M'):
    if value is None:
        return ""
    if isinstance(value, timedelta):
        # Handle timedelta
        return str(value)
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            return value
    return value.strftime(format)


# duration filter for timedelta objects
@app.template_filter('duration')
def duration_filter(value):
    if value is None:
        return ""
    if isinstance(value, timedelta):
        hours = value.seconds // 3600
        minutes = (value.seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"
    return str(value)

@app.template_filter('currency')
def currency_filter(value, currency_code=None):
    """Format value as currency"""
    if value is None:
        return "£0.00"
    
    if currency_code is None:
        # Default behavior for other pages (unchanged)
        return f"£{value:,.2f}"
    
    # Handle different currencies for hotel listing
    symbols = {
        'GBP': '£',
        'USD': '$',
        'EUR': '€',
        'JPY': '¥',
        'AUD': 'A$',
        'CAD': 'C$',
        'CHF': 'Fr.',
        'CNY': '¥',
        'INR': '₹'
    }
    
    symbol = symbols.get(currency_code, currency_code + ' ')
    return f"{symbol}{value:,.2f}"


# Authentication Functions used in login
def verify_user(email, password):
    """Verify user credentials and email verification status"""
    try:
        cursor = get_db()
        cursor.execute('''
            SELECT u.*, r.role_name 
            FROM users u 
            JOIN user_roles r ON u.role_id = r.role_id 
            WHERE u.email = %s
        ''', (email,))
        user = cursor.fetchone()
        cursor.close()
        
        if user and check_password_hash(user['password_hash'], password):
            if not user['is_active']:
                return None, "Please verify your email before logging in"
            return user, None
        return None, "Invalid email or password"
    except Exception as e:
        app.logger.error(f'User verification error: {str(e)}')
        return None, "An error occurred during verification"
    
    
# Get wkhtmltopdf path for PDF generation
def get_wkhtmltopdf_path():
    if platform.system() == 'Windows':
        return r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
    else:
        return '/usr/local/bin/wkhtmltopdf'
    

# PDF generation function for booking confirmation
def generate_booking_pdf(booking_data):
    """Generate PDF for booking confirmation"""
    try:
        # Configure wkhtmltopdf path using the platform-specific function
        config = pdfkit.configuration(wkhtmltopdf=get_wkhtmltopdf_path())
        
        # Render HTML template with all necessary data
        html_content = render_template(
            'booking/pdf_template.html',
            booking=booking_data['booking'],
            primary_guest=booking_data['primary_guest'],
            additional_guests=booking_data['additional_guests'],
            cancellation_dates=booking_data['cancellation_dates'],
            qr_code_url=booking_data['qr_code_url']
        )
        
        # PDF options
        options = {
            'page-size': 'A4',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': "UTF-8",
            'no-outline': None,
            'enable-local-file-access': True
        }
        
        # Generate PDF from HTML
        pdf = pdfkit.from_string(html_content, False, options=options, configuration=config)
        return pdf
        
    except Exception as e:
        app.logger.error(f"PDF generation error: {str(e)}")
        raise Exception(f"Failed to generate PDF: {str(e)}")
    

# Helper function to get the appropriate icon class for an amenity
def get_amenity_icon(amenity_name):
    """Helper function to get the appropriate icon class for an amenity"""
    icon_mapping = {
        'WiFi': 'fa fa-wifi',
        'Parking': 'fa fa-parking',
        'Pool': 'fa fa-pool',
        'Gym': 'fa fa-gym',
        'Restaurant': 'fa fa-restaurant',
        'Bar': 'fa fa-bar',
        'Spa': 'fa fa-spa',
        'Room Service': 'fa fa-bell',
        'Air Conditioning': 'fa fa-thermometer',
        'Conference Room': 'fa fa-briefcase',
        'Business Center': 'fa fa-pc-display'
    }

    return icon_mapping.get(amenity_name, 'fa fa-check-circle')


# Inject now function for datetime
@app.context_processor
def inject_now():
    return {'now': datetime.utcnow()}


# Login route
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    form = LoginForm()
    if form.validate_on_submit():
        email = form.email.data
        password = form.password.data
        remember_me = form.remember_me.data
        
        try:
            user, error_message = verify_user(email, password)
            
            if user:
                # Store user info in session
                session['user_id'] = user['user_id']
                session['first_name'] = user['first_name']
                session['last_name'] = user['last_name']
                session['user_role'] = user['role_id']
                session['is_admin'] = (user['role_name'] == 'admin')

                # Set session to permanent if remember me is checked
                if remember_me:
                    session.permanent = True
                    app.permanent_session_lifetime = timedelta(days=30)
                else:
                    session.permanent = False
                
                # Update last login
                cursor = get_db()
                cursor.execute('UPDATE users SET last_login = NOW() WHERE user_id = %s', 
                             (user['user_id'],))
                mysql.connection.commit()
                cursor.close()
                
                flash('Login successful', 'success')

                if user['role_id'] == 1:  # Admin role
                    return redirect(url_for('admin_dashboard'))
                
                # Check for stored form token
                token = session.pop('form_token', None)
                next_url = session.pop('next_url', None)
                
                if token:
                    cursor = get_db()
                    cursor.execute('''
                        SELECT form_data 
                        FROM temp_form_data 
                        WHERE token = %s AND expiry > NOW()
                    ''', (token,))
                    result = cursor.fetchone()
                    cursor.close()
                    
                    if result:
                        cursor = get_db()
                        cursor.execute('DELETE FROM temp_form_data WHERE token = %s', (token,))
                        mysql.connection.commit()
                        cursor.close()
                        form_data = json.loads(result['form_data'])
                        return render_template('auth/secure_redirect.html',
                                            action=next_url,
                                            form_data=form_data)
                
                return redirect(next_url or url_for('index'))
            else:
                flash(error_message, 'error')
            
        except Exception as e:
            app.logger.error(f'Login error: {str(e)}')
            flash('An error occurred. Please try again.', 'error')
    
    return render_template('auth/login.html', form=form)


# clean up functions to automatically clean up the temporarily stored form data
def cleanup_expired_form_data():
    try:
        cursor = get_db()
        cursor.execute('DELETE FROM temp_form_data WHERE expiry < NOW()')
        mysql.connection.commit()
        app.logger.info('Cleaned up expired form data')
    except Exception as e:
        app.logger.error(f'Cleanup error: {str(e)}')
    finally:
        if cursor:
            cursor.close()

# Cleanup function for expired form data called periodically
@app.before_request
def before_request():
    # Run cleanup occasionally (e.g., 1% of requests)
    if random.random() < 0.01:  # 1% chance of running cleanup on each request
        cleanup_expired_form_data()


# Register route
@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("3 per hour")
def register():
    form = RegisterForm()
    if form.validate_on_submit():
        try:
            cursor = get_db()
            
            # Check if email exists
            cursor.execute('SELECT user_id FROM users WHERE email = %s', (form.email.data,))
            if cursor.fetchone():
                flash('Email already exists', 'error')
                return redirect(url_for('register'))

            # Generate verification token
            token = secrets.token_urlsafe(32)
            
            # Insert new user
            cursor.execute('''
                INSERT INTO users (
                    role_id, first_name, last_name, email, 
                    password_hash, phone_number, is_active,
                    created_at, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, FALSE, 
                    CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ''', (
                3,  # Customer role
                form.first_name.data,
                form.last_name.data,
                form.email.data,
                generate_password_hash(form.password.data),
                form.phone.data
            ))
            
            user_id = cursor.lastrowid
            
            # Store verification token
            cursor.execute('''
                INSERT INTO email_verification (user_id, token, expires_at)
                VALUES (%s, %s, DATE_ADD(NOW(), INTERVAL 24 HOUR))
            ''', (user_id, token))
            
            mysql.connection.commit()
            
            # Send verification email
            verification_url = url_for('verify_email', token=token, _external=True)
            if send_email(
                to=form.email.data,
                subject='Verify Your Email - World Hotels',
                template='auth/verify_email.html',
                verification_url=verification_url
            ):
                flash('Registration successful! Please check your email to verify your account.', 'success')
            else:
                flash('Registration successful but email verification failed. Please contact support.', 'warning')
            
            return redirect(url_for('login'))

        except Exception as e:
            if cursor:
                cursor.close()
            app.logger.error(f"Registration error: {str(e)}")
            flash('An error occurred. Please try again.', 'error')
            return redirect(url_for('register'))

    return render_template('auth/register.html', form=form)


# Email verification route after registration
@app.route('/verify-email/<token>')
def verify_email(token):
    try:
        cursor = get_db()
        
        # Check token validity
        cursor.execute('''
            SELECT user_id 
            FROM email_verification 
            WHERE token = %s AND expires_at > NOW() AND used = 0
        ''', (token,))
        verification = cursor.fetchone()
        
        if verification:
            # Activate user account
            cursor.execute('''
                UPDATE users 
                SET is_active = TRUE 
                WHERE user_id = %s
            ''', (verification['user_id'],))
            
            # Mark token as used
            cursor.execute('''
                UPDATE email_verification 
                SET used = 1 
                WHERE token = %s
            ''', (token,))
            
            mysql.connection.commit()
            flash('Email verified successfully! You can now login.', 'success')
        else:
            flash('Invalid or expired verification link.', 'error')
            
    except Exception as e:
        app.logger.error(f"Email verification error: {str(e)}")
        flash('An error occurred during verification.', 'error')
    finally:
        cursor.close()
        
    return redirect(url_for('login'))


# Forgot password route
@app.route('/forgot-password', methods=['GET', 'POST'])
@limiter.limit("20 per hour")
def forgot_password():
    form = ForgotPasswordForm()
    if form.validate_on_submit():
        email = form.email.data
        try:
            cursor = get_db()
            cursor.execute('SELECT user_id FROM users WHERE email = %s', (email,))
            user = cursor.fetchone()

            if user:
                token = secrets.token_urlsafe(32)
                expiry = datetime.now() + timedelta(hours=1)

                cursor.execute('''
                    INSERT INTO password_reset_tokens (user_id, token, expires_at)
                    VALUES (%s, %s, %s)
                ''', (user['user_id'], token, expiry))
                
                mysql.connection.commit()

                # Send reset email using the updated function
                if send_password_reset_email(email, token):
                    flash('Password reset instructions have been sent to your email.', 'success')
                else:
                    flash('Error sending reset email. Please try again.', 'error')
            else:
                # Don't reveal if email exists
                flash('If your email exists in our system, you will receive a reset link', 'info')

        except Exception as e:
            app.logger.error(f'Forgot password error: {str(e)}')
            flash('An error occurred. Please try again.', 'error')
        finally:
            if cursor:
                cursor.close()

    return render_template('auth/forgot_password.html', form=form)


# Reset password route
@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    form = ResetPasswordForm()
    try:
        cursor = get_db()
        # Get user info and current password hash
        cursor.execute('''
            SELECT t.user_id, u.password_hash 
            FROM password_reset_tokens t 
            JOIN users u ON t.user_id = u.user_id 
            WHERE t.token = %s AND t.expires_at > NOW() AND t.used = 0
        ''', (token,))
        token_data = cursor.fetchone()

        if not token_data:
            flash('Invalid or expired reset link', 'error')
            return redirect(url_for('login'))

        if form.validate_on_submit():
            # Check if new password is same as old password
            if check_password_hash(token_data['password_hash'], form.password.data):
                flash('New password cannot be the same as your old password', 'error')
                return render_template('auth/reset_password.html', form=form, token=token)

            # Update password and mark token as used
            hashed_password = generate_password_hash(form.password.data)
            cursor.execute('UPDATE users SET password_hash = %s WHERE user_id = %s', 
                         (hashed_password, token_data['user_id']))
            cursor.execute('UPDATE password_reset_tokens SET used = 1 WHERE token = %s', 
                         (token,))
            
            mysql.connection.commit()
            
            # Log the password reset for security
            app.logger.info(f"Password reset successful for user_id: {token_data['user_id']}")
            
            flash('Your password has been reset successfully. Please login with your new password.', 'success')
            return redirect(url_for('login'))

    except Exception as e:
        app.logger.error(f'Reset password error: {str(e)}')
        flash('An error occurred. Please try again.', 'error')
    finally:
        if cursor:
            cursor.close()

    return render_template('auth/reset_password.html', form=form, token=token)


# Logout route
@app.route('/logout')
def logout():
    response = redirect(url_for('index'))
    
    # Clear remember-me token if exists
    remember_token = request.cookies.get('remember_token')
    if remember_token:
        try:
            selector = remember_token.split(':')[0]
            cursor = get_db()
            cursor.execute('DELETE FROM auth_tokens WHERE selector = %s', (selector,))
            mysql.connection.commit()
        except:
            pass
        response.delete_cookie('remember_token')
    
    session.clear()
    flash('You have been logged out', 'success')
    return response



# home page route
@app.route('/')
def index():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Fetch 6 random hotels with their basic info
        hotels_query = """
        SELECT 
            hotel_id,
            city,
            title,
            description,
            hotel_image
        FROM hotels
        WHERE status = 'active'
        AND hotel_image IS NOT NULL
        ORDER BY RAND()
        LIMIT 6
        """
        
        cursor.execute(hotels_query)
        destinations = cursor.fetchall()
        
        # Fetch gallery images
        gallery_query = """
            SELECT gallery_id, image_url, description 
            FROM gallery 
            WHERE is_active = TRUE 
            ORDER BY display_order ASC, gallery_id DESC
            LIMIT 11
        """
        cursor.execute(gallery_query)
        gallery_images = cursor.fetchall()
        
        cursor.close()
        
        # Create form for CSRF protection
        form = FlaskForm()
        
        return render_template('index.html', 
                             destinations=destinations,
                             gallery_images=gallery_images,
                             form=form)
        
    except Exception as e:
        app.logger.error(f"Error fetching data: {str(e)}")
        return render_template('index.html', 
                             destinations=[],
                             gallery_images=[],
                             form=FlaskForm())
    

# Initialize serializer with your app secret key
def get_token_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'])

@app.route('/subscribe-newsletter', methods=['POST'])
@limiter.limit("5 per minute")
def subscribe_newsletter():
    try:
        email = request.form.get('email')
        if not email:
            return jsonify({'success': False, 'message': 'Email is required'}), 400

        cursor = get_db()
        
        # Check if email already exists
        cursor.execute('SELECT subscriber_id, is_active FROM newsletter_subscribers WHERE email = %s', (email,))
        existing_subscriber = cursor.fetchone()
        
        if existing_subscriber:
            if existing_subscriber['is_active']:
                return jsonify({'success': False, 'message': 'You are already subscribed!'}), 400
            else:
                # Reactivate subscription
                cursor.execute('''
                    UPDATE newsletter_subscribers 
                    SET is_active = TRUE, updated_at = NOW() 
                    WHERE email = %s
                ''', (email,))
        else:
            # Add new subscriber
            cursor.execute('''
                INSERT INTO newsletter_subscribers (email, is_active)
                VALUES (%s, TRUE)
            ''', (email,))
        
        mysql.connection.commit()

        # Generate unsubscribe token
        s = get_token_serializer()
        unsubscribe_token = s.dumps(email, salt='unsubscribe-salt')

        # Send welcome email
        send_email(
            to=email,
            subject='Welcome to World Hotels Newsletter',
            template='email/newsletter_welcome.html',
            unsubscribe_token=unsubscribe_token
        )

        return jsonify({
            'success': True, 
            'message': 'Thank you for subscribing to our newsletter!'
        })

    except Exception as e:
        app.logger.error(f"Newsletter subscription error: {str(e)}")
        return jsonify({
            'success': False, 
            'message': 'An error occurred. Please try again.'
        }), 500
    finally:
        if cursor:
            cursor.close()

# Newsletter unsubscribe route
@app.route('/unsubscribe/<token>')
def unsubscribe_newsletter(token):
    try:
        s = get_token_serializer()
        email = s.loads(token, salt='unsubscribe-salt', max_age=2592000)  # 30 days expiry
        
        cursor = get_db()
        cursor.execute('''
            UPDATE newsletter_subscribers 
            SET is_active = FALSE, updated_at = NOW() 
            WHERE email = %s
        ''', (email,))
        
        mysql.connection.commit()
        cursor.close()
        
        return render_template('email/unsubscribe.html', success=True)
        
    except Exception as e:
        app.logger.error(f"Newsletter unsubscribe error: {str(e)}")
        return render_template('email/unsubscribe.html', success=False)


# Booking search route handle search page
@app.route('/search')
def search():
    try:
        cursor = get_db()
        
        # Fetch all active cities from hotels
        cursor.execute('''
            SELECT DISTINCT 
                h.city,
                h.hotel_id,
                MIN(r.base_price) as standard_room_price,
                COUNT(DISTINCT r.room_id) as capacity
            FROM hotels h
            JOIN rooms r ON h.hotel_id = r.hotel_id
            WHERE h.status = 'active'
            AND r.is_active = TRUE
            GROUP BY h.city, h.hotel_id
            ORDER BY h.city
        ''')
        cities = cursor.fetchall()

        # Fetch recommended hotels (based on capacity and availability)
        cursor.execute('''
            SELECT 
                h.hotel_id,
                h.hotel_name,
                h.city,
                h.hotel_image,
                h.description,
                MIN(r.base_price) as starting_price,
                COUNT(DISTINCT r.room_id) as available_rooms
            FROM hotels h
            JOIN rooms r ON h.hotel_id = r.hotel_id
            WHERE h.status = 'active'
            AND r.status = 'available'
            GROUP BY h.hotel_id, h.hotel_name, h.city, h.hotel_image, h.description
            ORDER BY available_rooms DESC, starting_price ASC
            LIMIT 6
        ''')
        recommended_hotels = cursor.fetchall() # Fetch recommended hotels

        # Fetch currencies
        cursor.execute('SELECT * FROM currencies WHERE is_active = TRUE')
        currencies = cursor.fetchall()

        cursor.close()

        return render_template('booking/search.html',
                             cities=cities,
                             recommended_hotels=recommended_hotels,
                             currencies=currencies,
                             selected_currency=session.get('currency', 'GBP'))

    # Exception handling for search page
    except Exception as e:
        app.logger.error(f'Search page error: {str(e)}')
        flash('An error occurred while loading the search page', 'error')
        return redirect(url_for('index'))
    

# Hotel list route handle hotel list page
@app.route('/hotel-list')
def hotel_list():
    try:
        # Get search parameters
        city = request.args.get('city')
        check_in = datetime.strptime(request.args.get('check_in'), '%Y-%m-%d')
        check_out = datetime.strptime(request.args.get('check_out'), '%Y-%m-%d')
        adults = int(request.args.get('adults', 1))
        children = int(request.args.get('children', 0))
        rooms = int(request.args.get('rooms', 1))
        currency = request.args.get('currency', 'GBP')
        sort_by = request.args.get('sort', 'price_asc')

        total_guests = adults + children
        total_nights = (check_out - check_in).days
        guests_per_room = math.ceil(total_guests / rooms)

        cursor = get_db()

        # Updated SQL query compatible with ONLY_FULL_GROUP_BY
        # Modify the SQL query to handle 'all' cities
        base_query = '''
            SELECT 
                h.*,
                MIN(r.base_price) as min_price,
                MAX(r.base_price) as max_price,
                COUNT(DISTINCT r.room_id) as total_rooms,
                GROUP_CONCAT(DISTINCT rt.type_name) as room_types
            FROM hotels h
            JOIN rooms r ON h.hotel_id = r.hotel_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            WHERE h.status = 'active'
            AND r.status = 'available'
            AND r.is_active = TRUE
            {}
            AND EXISTS (
                SELECT 1 
                FROM rooms r2 
                JOIN room_types rt2 ON r2.room_type_id = rt2.room_type_id
                WHERE r2.hotel_id = h.hotel_id
                AND r2.status = 'available'
                AND r2.is_active = TRUE
                AND rt2.max_occupancy >= %s
            )
            GROUP BY h.hotel_id
            HAVING total_rooms >= %s
        '''

        # Check if city is not 'all'
        if city and city.lower() != 'all':
            where_clause = 'AND h.city = %s'
            params = (city, guests_per_room, rooms)
        else:
            where_clause = ''
            params = (guests_per_room, rooms)

        cursor.execute(base_query.format(where_clause), params) # Execute query
        
        hotels = [] # Initialize hotels list
        for hotel_data in cursor.fetchall():
            hotel = dict(hotel_data)
            
            # Get available rooms by type
            cursor.execute('''
                SELECT 
                    rt.type_name,
                    rt.room_type_id,
                    COUNT(r.room_id) as available,
                    MIN(r.base_price) as base_price,
                    rt.max_occupancy,
                    rt.description as room_description
                FROM rooms r
                JOIN room_types rt ON r.room_type_id = rt.room_type_id
                WHERE r.hotel_id = %s
                AND r.status = 'available'
                AND r.is_active = TRUE
                GROUP BY rt.type_name, rt.room_type_id, rt.max_occupancy, rt.description
            ''', (hotel['hotel_id'],))
            hotel['available_rooms'] = cursor.fetchall()

            # Get hotel amenities 
            cursor.execute('''
                SELECT DISTINCT
                    MIN(ha.amenity_id) as amenity_id,
                    ha.amenity_name as name,
                    ha.description,
                    ha.icon_class
                FROM hotel_amenities ha
                WHERE ha.hotel_id = %s
                GROUP BY ha.amenity_name, ha.description, ha.icon_class
            ''', (hotel['hotel_id'],))
            hotel['amenities'] = cursor.fetchall()

            # Calculate pricing
            is_peak = check_in.month in [4, 5, 6, 7, 8, 11, 12]
            base_price = float(hotel['min_price']) * (1.2 if is_peak else 1)
            
            # Apply advance booking discount
            days_in_advance = (check_in - datetime.now()).days
            if days_in_advance >= 80:
                hotel['discount_percentage'] = 30
            elif days_in_advance >= 60:
                hotel['discount_percentage'] = 20
            elif days_in_advance >= 45:
                hotel['discount_percentage'] = 10
            else:
                hotel['discount_percentage'] = 0

            # Calculate total price
            hotel['price'] = base_price * (1 - hotel['discount_percentage'] / 100)
            hotel['total_price'] = hotel['price'] * total_nights

            # Convert currency if needed
            if currency != 'GBP':
                cursor.execute('''
                    SELECT rate
                    FROM exchange_rates 
                    WHERE from_currency_id = (SELECT currency_id FROM currencies WHERE code = 'GBP')
                    AND to_currency_id = (SELECT currency_id FROM currencies WHERE code = %s)
                ''', (currency,))
                rate = float(cursor.fetchone()['rate'])
                hotel['price'] *= rate
                hotel['total_price'] *= rate
            
            hotels.append(hotel)

        # Get room types and amenities for filters
        cursor.execute('''
            SELECT DISTINCT 
                rt.room_type_id,
                rt.type_name,
                rt.description,
                rt.max_occupancy
            FROM room_types rt
            JOIN rooms r ON rt.room_type_id = r.room_type_id
            WHERE r.is_active = TRUE
            ORDER BY rt.max_occupancy
        ''')
        room_types = cursor.fetchall()

        # Get amenities for filters
        cursor.execute('''
            SELECT DISTINCT 
                MIN(ha.amenity_id) as amenity_id,
                ha.amenity_name as name,
                ha.description,
                ha.icon_class
            FROM hotel_amenities ha
            GROUP BY ha.amenity_name, ha.description, ha.icon_class
            ORDER BY ha.amenity_name
        ''')
        amenities = cursor.fetchall()

        # Apply sorting
        if sort_by == 'price_asc':
            hotels.sort(key=lambda x: x['price'])
        elif sort_by == 'price_desc':
            hotels.sort(key=lambda x: x['price'], reverse=True)
        elif sort_by == 'name_asc':
            hotels.sort(key=lambda x: x['hotel_name'])

        cursor.close()

        # Calculate price range for filter
        min_price = min([h['price'] for h in hotels]) if hotels else 0
        max_price = max([h['price'] for h in hotels]) if hotels else 0

        return render_template('booking/hotel_list.html',
                             hotels=hotels,
                             room_types=room_types,
                             amenities=amenities,
                             city=city,
                             check_in=check_in,
                             check_out=check_out,
                             total_nights=total_nights,
                             total_guests=total_guests,
                             rooms=rooms,
                             selected_currency=currency,
                             min_price=float(min_price),
                             max_price=float(max_price),
                             sort_by=sort_by)

    except Exception as e:
        app.logger.error(f"Hotel list error: {str(e)}")
        flash('An error occurred while loading hotels', 'error')
        return redirect(url_for('search'))


# Room list route handle room list page after hotel list page
@app.route('/room-list')
def room_list():
    try:
        # Check required parameters
        if not all([
            request.args.get('hotel_id'),
            request.args.get('check_in'),
            request.args.get('check_out')
        ]):
            flash('Missing required parameters', 'error')
            return redirect(url_for('search'))

        # Get parameters from request
        hotel_id = request.args.get('hotel_id')
        check_in = datetime.strptime(request.args.get('check_in'), '%Y-%m-%d')
        check_out = datetime.strptime(request.args.get('check_out'), '%Y-%m-%d')
        total_guests = int(request.args.get('guests', 1))
        currency = request.args.get('currency', 'GBP')

        total_nights = (check_out - check_in).days # Calculate total nights

        cursor = get_db()

        # Get maximum room capacity for this hotel
        cursor.execute('''
            SELECT MAX(rt.max_occupancy) as max_capacity
            FROM rooms r
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            WHERE r.hotel_id = %s
            AND r.status = 'available'
        ''', (hotel_id,))
        max_room_capacity = cursor.fetchone()['max_capacity'] # Fetch maximum room capacity

        # Get available currencies
        cursor.execute('SELECT * FROM currencies WHERE is_active = 1')
        currencies = cursor.fetchall() # Fetch available currencies

        # Get hotel details with amenities
        cursor.execute('''
            SELECT h.*, 
                   GROUP_CONCAT(ha.amenity_name) as amenities
            FROM hotels h
            LEFT JOIN hotel_amenities ha ON h.hotel_id = ha.hotel_id
            WHERE h.hotel_id = %s
            GROUP BY h.hotel_id
        ''', (hotel_id,))
        hotel = cursor.fetchone() # Fetch hotel details

        # Check if hotel is found
        if not hotel:
            flash('Hotel not found', 'error')
            return redirect(url_for('search'))

        # Process amenities into a list
        if hotel['amenities']:
            hotel['amenities'] = [
                {'amenity_name': name, 'icon_class': 'bi bi-check-circle'}  # Using a default icon
                for name in hotel['amenities'].split(',')
            ]
        else:
            hotel['amenities'] = []

        # Get available rooms with capacity filter
        cursor.execute('''
            SELECT 
                MIN(r.room_id) as room_id, 
                r.hotel_id, 
                r.room_type_id, 
                MIN(r.room_number) as room_number, 
                MIN(r.floor_number) as floor_number, 
                MIN(r.base_price) as base_price, 
                MIN(r.room_image) as room_image, 
                rt.type_name, 
                rt.max_occupancy, 
                rt.base_price_multiplier,
                COUNT(DISTINCT r2.room_id) as available_rooms
            FROM rooms r
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            LEFT JOIN rooms r2 ON r.room_type_id = r2.room_type_id 
                AND r2.hotel_id = r.hotel_id
                AND r2.status = 'available'
            WHERE r.hotel_id = %s
            AND rt.max_occupancy >= %s
            AND r.status = 'available'
            GROUP BY r.hotel_id, r.room_type_id, rt.type_name, rt.max_occupancy, rt.base_price_multiplier
            HAVING available_rooms > 0
        ''', (hotel_id, total_guests))
        available_rooms = cursor.fetchall()

        # Check if no rooms are available
        if not available_rooms:
            flash('No rooms available for the selected dates', 'info')
            return redirect(url_for('search'))

        # Get room features using join
        for room in available_rooms:
            cursor.execute('''
                SELECT DISTINCT f.*
                FROM room_features f
                JOIN room_feature_mapping rfm ON f.feature_id = rfm.feature_id
                WHERE rfm.room_id = %s
            ''', (room['room_id'],))
            room['features'] = cursor.fetchall()

            # Convert Decimal to float for calculations
            base_price = float(room['base_price'])
            
            # Calculate pricing
            is_peak = check_in.month in [4, 5, 6, 7, 8, 11, 12]
            base_price = base_price if is_peak else base_price * 0.5
            
            # Apply room type multiplier
            base_price *= float(room['base_price_multiplier'])

            # Extra guest charges
            if room['type_name'] == 'Double' and total_guests == 2:
                room['extra_guest_charge'] = base_price * 0.1
            else:
                room['extra_guest_charge'] = 0

            # Calculate advance booking discount
            days_in_advance = (check_in - datetime.now()).days
            if 80 <= days_in_advance <= 90:
                discount_percentage = 30
            elif 60 <= days_in_advance <= 79:
                discount_percentage = 20
            elif 45 <= days_in_advance <= 59:
                discount_percentage = 10
            else:
                discount_percentage = 0

            # Store all prices as float
            room['base_price'] = base_price
            room['price'] = base_price + room['extra_guest_charge']
            room['discount_amount'] = (room['price'] * total_nights) * (discount_percentage / 100)
            room['total_price'] = (room['price'] * total_nights) - room['discount_amount']

            # Convert currency if needed
            if currency != 'GBP':
                cursor.execute('''
                    SELECT rate 
                    FROM exchange_rates 
                    WHERE from_currency_id = (SELECT currency_id FROM currencies WHERE code = 'GBP')
                    AND to_currency_id = (SELECT currency_id FROM currencies WHERE code = %s)
                ''', (currency,))
                rate = float(cursor.fetchone()['rate'])
                
                for price_field in ['base_price', 'extra_guest_charge', 'price', 'discount_amount', 'total_price']:
                    room[price_field] *= rate

        cursor.close() # Close cursor after use

        return render_template('booking/room_list.html',
                             hotel=hotel,
                             available_rooms=available_rooms,
                             check_in=check_in,
                             check_out=check_out,
                             total_nights=total_nights,
                             total_guests=total_guests,
                             selected_currency=currency,
                             discount_percentage=discount_percentage,
                             max_room_capacity=max_room_capacity,
                             currencies=currencies,
                             days_in_advance=days_in_advance)

    except Exception as e:
        app.logger.error(f'Room list error: {str(e)}')
        flash('An error occurred while loading rooms', 'error')
        return redirect(url_for('search'))


# Booking form route handle booking form page
@app.route('/booking-form', methods=['GET', 'POST'])
@login_required
def booking_form():
    try:
        if request.method == 'POST':
            # Store form data in session
            session['booking_data'] = {
                'hotel_id': request.form.get('hotel_id'),
                'room_id': request.form.get('room_id'),
                'check_in': request.form.get('check_in'),
                'check_out': request.form.get('check_out'),
                'guests': request.form.get('guests', '1'),
                'currency': request.form.get('currency', 'GBP')
            }
            # Redirect to GET request
            return redirect(url_for('booking_form'))

        # Handle GET request
        params = session.get('booking_data', {}) if request.method == 'GET' else request.form

        # Parameter validation
        if not all([
            params.get('hotel_id'),
            params.get('room_id'),
            params.get('check_in'),
            params.get('check_out')
        ]):
            flash('Missing required parameters', 'error')
            return redirect(url_for('search'))

        # Process parameters
        hotel_id = params.get('hotel_id')
        room_id = params.get('room_id')
        check_in_str = params.get('check_in').split()[0]
        check_out_str = params.get('check_out').split()[0]
        check_in = datetime.strptime(check_in_str, '%Y-%m-%d')
        check_out = datetime.strptime(check_out_str, '%Y-%m-%d')
        total_guests = int(params.get('guests', 1))
        currency = params.get('currency', 'GBP')

        # Rest of your existing code remains the same...
        total_nights = (check_out - check_in).days
        cursor = get_db()
        

        # Get hotel and room details
        cursor.execute('SELECT * FROM hotels WHERE hotel_id = %s', (hotel_id,))
        hotel = cursor.fetchone()

        # Get room details
        cursor.execute('''
            SELECT r.*, rt.type_name, rt.max_occupancy, rt.base_price_multiplier
            FROM rooms r
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            WHERE r.room_id = %s
        ''', (room_id,))
        room = cursor.fetchone()

        # Convert Decimal to float for calculations
        base_price = float(room['base_price'])
        base_price_multiplier = float(room['base_price_multiplier'])

        # Calculate pricing
        is_peak = check_in.month in [4, 5, 6, 7, 8, 11, 12]
        base_price = base_price if is_peak else base_price * 0.5
        
        # Apply room type multiplier
        base_price *= base_price_multiplier

        # Extra guest charges
        extra_guest_charge = base_price * 0.1 if room['type_name'] == 'Double' and total_guests == 2 else 0

        # Calculate advance booking discount
        days_in_advance = (check_in - datetime.now()).days
        if 80 <= days_in_advance <= 90:
            discount_percentage = 30
        elif 60 <= days_in_advance <= 79:
            discount_percentage = 20
        elif 45 <= days_in_advance <= 59:
            discount_percentage = 10
        else:
            discount_percentage = 0

        # Calculate totals
        subtotal = (base_price + extra_guest_charge) * total_nights
        discount_amount = subtotal * (discount_percentage / 100)
        total_price = subtotal - discount_amount

        # Convert currency if needed
        if currency != 'GBP':
            cursor.execute('''
                SELECT CAST(rate AS DECIMAL(10,4)) as rate
                FROM exchange_rates 
                WHERE from_currency_id = (SELECT currency_id FROM currencies WHERE code = 'GBP')
                AND to_currency_id = (SELECT currency_id FROM currencies WHERE code = %s)
            ''', (currency,))
            rate = float(cursor.fetchone()['rate'])
            
            # Convert prices to new currency
            base_price *= rate
            extra_guest_charge *= rate
            subtotal *= rate
            discount_amount *= rate
            total_price *= rate

        # Store calculated prices in room dict
        room['base_price'] = base_price
        room['extra_guest_charge'] = extra_guest_charge

        cursor.close()

        return render_template('booking/booking_form.html',
                             hotel=hotel,
                             room=room,
                             check_in=check_in,
                             check_out=check_out,
                             total_nights=total_nights,
                             total_guests=total_guests,
                             selected_currency=currency,
                             subtotal=subtotal,
                             discount_percentage=discount_percentage,
                             discount_amount=discount_amount,
                             total_price=total_price)

    except Exception as e:
        app.logger.error(f'Booking form error: {str(e)}')
        flash('An error occurred while loading the booking form', 'error')
        return redirect(url_for('room_list'))
    

# Process booking route handle booking process
@app.route('/process-booking', methods=['POST'])
@login_required
def process_booking():
    try:
        # Check for CSRF token
        if not request.form.get('csrf_token'):
            abort(400, 'CSRF token missing')
            
        cursor = get_db()
        
        # Get form data
        hotel_id = request.form.get('hotel_id')
        room_id = request.form.get('room_id')
        check_in = datetime.strptime(request.form.get('check_in'), '%Y-%m-%d')
        check_out = datetime.strptime(request.form.get('check_out'), '%Y-%m-%d')
        total_amount = float(request.form.get('total_amount'))
        currency = request.form.get('currency')
        total_guests = int(request.form.get('total_guests', 1))
        
        # Generate booking reference
        booking_reference = f"WH-{datetime.now().year}-{random.randint(10000, 99999)}"
        
        # Insert booking record
        cursor.execute('''
            INSERT INTO bookings (
                booking_reference, user_id, hotel_id, room_id, currency_id,
                check_in_date, check_out_date, number_of_guests,
                total_amount, final_amount, booking_date, status, payment_status
            ) VALUES (
                %s, %s, %s, %s, 
                (SELECT currency_id FROM currencies WHERE code = %s),
                %s, %s, %s, %s, %s, NOW(), 'confirmed', 'paid'
            )
        ''', (
            booking_reference, 
            session['user_id'],
            hotel_id, 
            room_id,
            currency,
            check_in,
            check_out,
            total_guests,
            total_amount,
            total_amount  # final_amount is same as total_amount initially
        ))
        
        # Get the booking_id
        cursor.execute('SELECT LAST_INSERT_ID() as booking_id')
        booking_id = cursor.fetchone()['booking_id']
        
        # Insert primary guest details
        cursor.execute('''
            INSERT INTO booking_details (
                booking_id, guest_name, guest_email, 
                guest_phone, is_primary_guest
            ) VALUES (%s, %s, %s, %s, true)
        ''', (
            booking_id,
            request.form.get('primary_guest_name'),
            request.form.get('primary_guest_email'),
            request.form.get('primary_guest_phone')
        ))
        
        # Insert additional guests
        for i in range(1, total_guests):
            if request.form.get(f'guest_name_{i}'):
                cursor.execute('''
                    INSERT INTO booking_details (
                        booking_id, guest_name, guest_email, 
                        is_primary_guest
                    ) VALUES (%s, %s, %s, false)
                ''', (
                    booking_id,
                    request.form.get(f'guest_name_{i}'),
                    request.form.get(f'guest_email_{i}')
                ))
        
        # Save special requests if any
        if request.form.get('special_requests'):
            cursor.execute('''
                UPDATE bookings 
                SET special_requests = %s 
                WHERE booking_id = %s
            ''', (request.form.get('special_requests'), booking_id))
        
        # Insert transaction record
        cursor.execute('''
            INSERT INTO booking_transactions (
                booking_id, amount, transaction_type, 
                status, payment_method
            ) VALUES (%s, %s, 'payment', 'completed', 'credit_card')
        ''', (booking_id, total_amount))
        
        cursor.connection.commit()
        cursor.close()
        
        try:
            # Fetch booking details including hotel and room info
            cursor = get_db()
            cursor.execute('''
                SELECT 
                    b.*, h.*, r.*, rt.type_name,
                    c.code as currency_code,
                    c.symbol as currency_symbol,
                    DATEDIFF(b.check_out_date, b.check_in_date) as total_nights
                FROM bookings b
                JOIN hotels h ON b.hotel_id = h.hotel_id
                JOIN rooms r ON b.room_id = r.room_id
                JOIN room_types rt ON r.room_type_id = rt.room_type_id
                JOIN currencies c ON b.currency_id = c.currency_id
                WHERE b.booking_reference = %s
            ''', (booking_reference,))
            booking = cursor.fetchone()

            # Fetch guest details
            cursor.execute('''
                SELECT * FROM booking_details
                WHERE booking_id = %s
                ORDER BY is_primary_guest DESC
            ''', (booking_id,))
            guests = cursor.fetchall()

            primary_guest = next((g for g in guests if g['is_primary_guest']), None)
            additional_guests = [g for g in guests if not g['is_primary_guest']]

            # Calculate cancellation dates
            check_in = booking['check_in_date']
            cancellation_dates = {
                'free': check_in - timedelta(days=60),
                'partial': check_in - timedelta(days=30)
            }

            # Generate QR code URL
            qr_code_url = f"https://api.qrserver.com/v1/create-qr-code/?size=150x150&data={request.host_url}booking/{booking_reference}"

            # Generate PDF
            pdf_content = generate_booking_pdf({
                'booking': booking,
                'primary_guest': primary_guest,
                'additional_guests': additional_guests,
                'cancellation_dates': cancellation_dates,
                'qr_code_url': qr_code_url
            })

            # Create and send email
            msg = Message(
                subject=f'Booking Confirmation - {booking_reference}',
                recipients=[primary_guest['guest_email']],
                html=render_template(
                    'email/booking_confirmation_email.html',
                    booking=booking,
                    primary_guest=primary_guest,
                    additional_guests=additional_guests,
                    cancellation_dates=cancellation_dates,
                    qr_code_url=qr_code_url
                )
            )

            # Attach PDF to email
            msg.attach(
                f"booking-{booking_reference}.pdf",
                "application/pdf",
                pdf_content
            )

            # Send email
            mail.send(msg)
            cursor.close()

        except Exception as email_error:
            app.logger.error(f"Email sending error: {str(email_error)}")
            flash('Booking confirmed but confirmation email could not be sent', 'warning')
        
        # Clean up session data after successful booking
        session.pop('booking_data', None)
        
        # Redirect to confirmation page
        return redirect(url_for('booking_confirmation', 
                              booking_reference=booking_reference))
        
    except Exception as e:
        if cursor:
            cursor.connection.rollback()
        app.logger.error(f"Booking error: {str(e)}\n{traceback.format_exc()}")
        flash('An error occurred while processing your booking', 'error')
        return redirect(url_for('booking_form'))
    

# Booking confirmation route handle booking confirmation page with details and booking reference
@app.route('/booking-confirmation/<booking_reference>')
@login_required
def booking_confirmation(booking_reference):
    try:
        cursor = get_db()
        
        # Get booking details
        cursor.execute('''
            SELECT 
                b.*,
                h.*,
                r.*,
                rt.type_name,
                c.code as currency_code,
                c.symbol as currency_symbol,
                DATEDIFF(b.check_out_date, b.check_in_date) as total_nights
            FROM bookings b
            JOIN hotels h ON b.hotel_id = h.hotel_id
            JOIN rooms r ON b.room_id = r.room_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            JOIN currencies c ON b.currency_id = c.currency_id
            WHERE b.booking_reference = %s
            AND b.user_id = %s
        ''', (booking_reference, session['user_id']))
        
        booking = cursor.fetchone()
        
        if not booking:
            flash('Booking not found', 'error')
            return redirect(url_for('my_bookings'))

        # Get guest details from booking_details table
        cursor.execute('''
            SELECT * FROM booking_details
            WHERE booking_id = %s
            ORDER BY is_primary_guest DESC
        ''', (booking['booking_id'],))
        guests = cursor.fetchall()

        primary_guest = next((g for g in guests if g['is_primary_guest']), None)
        additional_guests = [g for g in guests if not g['is_primary_guest']]

        # Calculate cancellation dates
        check_in = booking['check_in_date']
        cancellation_dates = {
            'free': check_in - timedelta(days=60),
            'partial': check_in - timedelta(days=30)
        }

        # Generate QR code URL
        qr_code_url = f"https://api.qrserver.com/v1/create-qr-code/?size=150x150&data={request.host_url}booking/{booking_reference}"
        
        cursor.close()

        return render_template('booking/booking_confirmation.html',
                             booking=booking,
                             primary_guest=primary_guest,
                             additional_guests=additional_guests,
                             cancellation_dates=cancellation_dates,
                             qr_code_url=qr_code_url)

    except Exception as e:
        app.logger.error(f'Booking confirmation error: {str(e)}')
        flash('An error occurred while loading the confirmation', 'error')
        return redirect(url_for('my_bookings'))
    

# Verify booking route handle booking verification page with details and booking reference
@app.route('/booking/<booking_reference>')
def verify_booking(booking_reference):
    try:
        cursor = get_db()
        
        # Get booking details without requiring login
        cursor.execute('''
            SELECT 
                b.*,
                h.hotel_name,
                h.address,
                rt.type_name,
                bd.guest_name as primary_guest_name,
                b.check_in_date,
                b.check_out_date,
                b.status
            FROM bookings b
            JOIN hotels h ON b.hotel_id = h.hotel_id
            JOIN rooms r ON b.room_id = r.room_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            JOIN booking_details bd ON b.booking_id = bd.booking_id
            WHERE b.booking_reference = %s
            AND bd.is_primary_guest = true
        ''', (booking_reference,))
        
        booking = cursor.fetchone()
        cursor.close()
        
        # Check if booking is found
        if not booking:
            flash('Invalid booking reference', 'error')
            return redirect(url_for('index'))

        return render_template('booking/verify_booking.html', booking=booking) # Render verify booking page

    except Exception as e:
        app.logger.error(f'Booking verification error: {str(e)}')
        flash('An error occurred while verifying the booking', 'error')
        return redirect(url_for('index'))


# Download PDF route handle PDF download with booking details and booking reference
@app.route('/download-pdf/<booking_reference>')
@login_required
def download_pdf(booking_reference):
    try:
        cursor = get_db()
        
        # Get booking details
        cursor.execute('''
            SELECT 
                b.*,
                h.hotel_name,
                h.address,
                h.contact_number,
                h.email,
                h.check_in_time,
                h.check_out_time,
                rt.type_name,
                c.code as currency_code,
                c.symbol as currency_symbol,
                DATEDIFF(b.check_out_date, b.check_in_date) as total_nights
            FROM bookings b
            JOIN hotels h ON b.hotel_id = h.hotel_id
            JOIN rooms r ON b.room_id = r.room_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            JOIN currencies c ON b.currency_id = c.currency_id
            WHERE b.booking_reference = %s 
            AND b.user_id = %s
        ''', (booking_reference, session['user_id']))
        
        booking = cursor.fetchone()
        
        if not booking:
            flash('Booking not found', 'error')
            return redirect(url_for('my_bookings'))

        # Get guest details
        cursor.execute('''
            SELECT * FROM booking_details
            WHERE booking_id = %s
            ORDER BY is_primary_guest DESC
        ''', (booking['booking_id'],))
        guests = cursor.fetchall()

        primary_guest = next((g for g in guests if g['is_primary_guest']), None)
        additional_guests = [g for g in guests if not g['is_primary_guest']]

        # Calculate cancellation dates
        check_in = booking['check_in_date']
        cancellation_dates = {
            'free': check_in - timedelta(days=60),
            'partial': check_in - timedelta(days=30)
        }

        # Generate QR code URL
        qr_code_url = f"https://api.qrserver.com/v1/create-qr-code/?size=150x150&data={booking_reference}"

        # Configure wkhtmltopdf
        wkhtmltopdf_path = get_wkhtmltopdf_path()
        config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)
        

        # Generate HTML
        html = render_template('booking/pdf_template.html',
                             booking=booking,
                             primary_guest=primary_guest,
                             additional_guests=additional_guests,
                             cancellation_dates=cancellation_dates,
                             qr_code_url=qr_code_url)

        # PDF options
        options = {
            'page-size': 'A4',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': 'UTF-8',
            'no-outline': None,
            'enable-local-file-access': True,
            'quiet': None
        }

        try:
            # Generate PDF
            pdf = pdfkit.from_string(html, False, options=options, configuration=config)
            
            # Create response
            response = make_response(pdf)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename=booking-{booking_reference}.pdf'
            
            cursor.close()
            return response

        except Exception as pdf_error:
            app.logger.error(f'PDF generation specific error: {str(pdf_error)}')
            raise

    except Exception as e:
        app.logger.error(f'PDF generation error: {str(e)}')
        flash('Error generating PDF', 'error')
        return redirect(url_for('booking_confirmation', 
                              booking_reference=booking_reference))


# My bookings route handle my bookings page with filter and pagination and all bookings modifcations
@app.route('/my-bookings')
@login_required
def my_bookings():
    try:
        # Get filter parameters
        status = request.args.get('status')
        date_range = request.args.get('date_range')
        sort = request.args.get('sort', 'date_desc')
        page = int(request.args.get('page', 1))
        per_page = 10

        cursor = get_db()
        
        # Base query for counting
        count_query = '''
            SELECT COUNT(*) as total
            FROM bookings b
            WHERE b.user_id = %s
        '''
        count_params = [session['user_id']]

        # Apply filters to count query
        if status:
            count_query += ' AND b.status = %s'
            count_params.append(status)

        if date_range == 'upcoming':
            count_query += ' AND b.check_in_date >= CURDATE()'
        elif date_range == 'past':
            count_query += ' AND b.check_out_date < CURDATE()'

        # Get total count
        cursor.execute(count_query, count_params)
        total_count = cursor.fetchone()['total']
        pages = (total_count + per_page - 1) // per_page

        # Main query for bookings
        query = '''
            SELECT 
                b.*,
                h.hotel_name,
                h.hotel_image,
                rt.type_name as room_type,
                c.code as currency_code,  # Added currency code
                c.symbol as currency_symbol  # Added currency symbol
            FROM bookings b
            JOIN hotels h ON b.hotel_id = h.hotel_id
            JOIN rooms r ON b.room_id = r.room_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            JOIN currencies c ON b.currency_id = c.currency_id  # Added JOIN with currencies table
            WHERE b.user_id = %s
        '''
        params = [session['user_id']]

        # Apply filters
        if status:
            query += ' AND b.status = %s'
            params.append(status)

        if date_range == 'upcoming':
            query += ' AND b.check_in_date >= CURDATE()'
        elif date_range == 'past':
            query += ' AND b.check_out_date < CURDATE()'

        # Apply sorting
        if sort == 'date_asc':
            query += ' ORDER BY b.booking_date ASC'
        else:
            query += ' ORDER BY b.booking_date DESC'

        # Add pagination
        query += ' LIMIT %s OFFSET %s'
        params.extend([per_page, (page - 1) * per_page])

        # Get bookings
        cursor.execute(query, params)
        bookings = cursor.fetchall()

        # Process each booking
        for booking in bookings:
            # Check if booking can be cancelled
            days_until_checkin = (booking['check_in_date'] - datetime.now().date()).days
            booking['can_cancel'] = days_until_checkin > 0

            # Calculate cancellation charge if applicable
            if days_until_checkin <= 30:
                booking['cancellation_charge'] = booking['final_amount']
            elif days_until_checkin <= 60:
                # Convert float to Decimal for multiplication
                booking['cancellation_charge'] = booking['final_amount'] * Decimal('0.5')
            else:
                booking['cancellation_charge'] = Decimal('0')

            # Add formatted currency amounts
            booking['formatted_final_amount'] = f"{booking['currency_symbol']}{booking['final_amount']:,.2f}"
            booking['formatted_cancellation_charge'] = f"{booking['currency_symbol']}{booking['cancellation_charge']:,.2f}"

        cursor.close()

        return render_template('customer/my_bookings.html',
                             bookings=bookings,
                             status=status,
                             date_range=date_range,
                             sort=sort,
                             current_page=page,
                             pages=pages)

    except Exception as e:
        app.logger.error(f'My bookings error: {str(e)}')
        flash('An error occurred while loading your bookings', 'error')
        return redirect(url_for('index'))
    

# Cancel booking route handle booking cancellation for user with booking id
@app.route('/cancel-booking/<int:booking_id>', methods=['POST'])
@login_required
def cancel_booking(booking_id):
    try:
        cursor = get_db()
        
        # Check if booking exists and belongs to user
        cursor.execute('''
            SELECT * FROM bookings 
            WHERE booking_id = %s AND user_id = %s
        ''', (booking_id, session['user_id']))
        
        booking = cursor.fetchone()
        
        if not booking:
            flash('Booking not found', 'error')
            return redirect(url_for('my_bookings'))
            
        # Check if booking can be cancelled
        days_until_checkin = (booking['check_in_date'] - datetime.now().date()).days
        
        if days_until_checkin <= 0:
            flash('This booking cannot be cancelled', 'error')
            return redirect(url_for('my_bookings'))
            
        # Calculate cancellation charge
        if days_until_checkin <= 30:
            cancellation_charge = booking['final_amount']
        elif days_until_checkin <= 60:
            cancellation_charge = booking['final_amount'] * 0.5
        else:
            cancellation_charge = 0
            
        # Update booking status
        cursor.execute('''
            UPDATE bookings 
            SET status = 'cancelled',
                cancellation_charge = %s,
                updated_at = NOW()
            WHERE booking_id = %s
        ''', (cancellation_charge, booking_id))
        
        # Add status history record
        cursor.execute('''
            INSERT INTO booking_status_history 
            (booking_id, status, changed_by_user_id, notes)
            VALUES (%s, 'cancelled', %s, 'Cancelled by customer')
        ''', (booking_id, session['user_id']))
        
        # Commit changes
        cursor.connection.commit()
        cursor.close()
        
        flash('Booking cancelled successfully', 'success')
        return redirect(url_for('my_bookings'))
        
    except Exception as e:
        if cursor:
            cursor.connection.rollback()
        app.logger.error(f'Cancel booking error: {str(e)}')
        flash('An error occurred while cancelling the booking', 'error')
        return redirect(url_for('my_bookings'))


# Profile route handle user profile page with profile updates and password change
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    try:
        cursor = get_db()
        
        if request.method == 'POST':
            # Get and sanitize form data
            first_name = sanitize_input(request.form.get('first_name'))
            last_name = sanitize_input(request.form.get('last_name'))
            phone = sanitize_input(request.form.get('phone'))
            email = sanitize_input(request.form.get('email'))

            # Validate required fields
            if not all([first_name, last_name, email]):
                flash('Please fill in all required fields', 'error')
                return redirect(url_for('profile'))

            # Validate email format
            if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
                flash('Please enter a valid email address', 'error')
                return redirect(url_for('profile'))

            # Check if email is changed and already exists
            if email != session.get('user_email'):
                cursor.execute('SELECT user_id FROM users WHERE email = %s AND user_id != %s', 
                             (email, session['user_id']))
                if cursor.fetchone():
                    flash('Email address is already in use', 'error')
                    return redirect(url_for('profile'))

            try:
                # Update user details - removed full_name
                cursor.execute('''
                    UPDATE users 
                    SET first_name = %s,
                        last_name = %s,
                        email = %s,
                        phone_number = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE user_id = %s
                ''', (
                    first_name,
                    last_name,
                    email,
                    phone,
                    session['user_id']
                ))
                
                # Commit changes
                mysql.connection.commit()

                # Update session data
                session['user_email'] = email
                session['first_name'] = first_name
                session['last_name'] = last_name

                flash('Profile updated successfully', 'success')
                return redirect(url_for('profile'))

            except Exception as e:
                mysql.connection.rollback()
                app.logger.error(f"Profile update error: {str(e)}")
                flash('Error updating profile. Please try again.', 'error')
                return redirect(url_for('profile'))

        # GET request - fetch user profile
        cursor.execute('''
            SELECT 
                u.user_id,
                u.first_name,
                u.last_name,
                u.email,
                u.phone_number,
                u.created_at,
                u.last_login,
                COUNT(DISTINCT b.booking_id) as total_bookings,
                SUM(CASE WHEN b.status = 'confirmed' THEN 1 ELSE 0 END) as active_bookings,
                MAX(b.booking_date) as last_booking_date
            FROM users u
            LEFT JOIN bookings b ON u.user_id = b.user_id
            WHERE u.user_id = %s
            GROUP BY u.user_id
        ''', (session['user_id'],))
        
        user = cursor.fetchone()
        
        # Check if user profile is found
        if not user:
            flash('User profile not found', 'error')
            return redirect(url_for('logout'))

        # Add additional user stats
        user['member_days'] = (datetime.now() - user['created_at']).days
        user['is_active'] = user['last_login'] > (datetime.now() - timedelta(days=30)) if user['last_login'] else False
        
        return render_template('customer/profile.html', user=user)
                             
    except Exception as e:
        app.logger.error(f"Profile error: {str(e)}\n{traceback.format_exc()}")
        flash('An error occurred while accessing your profile', 'error')
        return redirect(url_for('index'))
    
    # Close cursor after use
    finally:
        if 'cursor' in locals():
            cursor.close()


# Change password route handle password change for user
@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    try:
        if request.method == 'POST':
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            cursor = get_db()
            
            # Check if user exists
            cursor.execute('SELECT password_hash FROM users WHERE user_id = %s', 
                         (session['user_id'],))
            user = cursor.fetchone()
            
            # Check if current password is correct
            if not check_password_hash(user['password_hash'], current_password):
                flash('Current password is incorrect', 'error')
                return redirect(url_for('change_password'))
            
            # Check if new passwords match
            if new_password != confirm_password:
                flash('New passwords do not match', 'error')
                return redirect(url_for('change_password'))
            
            # Check if new password is at least 8 characters long
            if len(new_password) < 8:
                flash('Password must be at least 8 characters long', 'error')
                return redirect(url_for('change_password'))
            
            # Generate password hash
            password_hash = generate_password_hash(new_password)
            
            # Update user password
            cursor.execute('''
                UPDATE users 
                SET password_hash = %s, updated_at = CURRENT_TIMESTAMP
                WHERE user_id = %s
            ''', (password_hash, session['user_id']))
            
            # Commit changes
            mysql.connection.commit()
            cursor.close()
            
            flash('Password updated successfully', 'success')
            return redirect(url_for('profile'))
            
        return render_template('customer/change_password.html')
        
    except Exception as e:
        app.logger.error(f'Password change error: {str(e)}')
        flash('Error changing password', 'error')
        return redirect(url_for('profile'))


# admin routes
# Admin dashboard route handle admin dashboard page with basic stats and recent bookings
@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    try:
        cursor = get_db()
        
         # Get basic stats with proper currency handling - FIXED QUERY
        cursor.execute('''
            SELECT 
                (SELECT COUNT(*) FROM bookings WHERE status = 'confirmed') as total_bookings,
                (SELECT COUNT(*) FROM hotels) as total_hotels,
                (SELECT COUNT(*) FROM users WHERE role_id = 3) as total_users,
                (SELECT SUM(CASE 
                    WHEN currency_id = 1 THEN final_amount  -- GBP
                    WHEN currency_id = 2 THEN final_amount * 0.79  -- USD to GBP
                    WHEN currency_id = 3 THEN final_amount * 0.85  -- EUR to GBP
                    ELSE final_amount
                END)
                FROM bookings 
                WHERE status = 'confirmed') as total_revenue
        ''')
        stats = cursor.fetchone()
        
        # Get recent bookings
        cursor.execute('''
            SELECT 
                b.booking_id,
                b.booking_reference,
                b.check_in_date,
                b.check_out_date,
                b.final_amount,
                b.status,
                h.hotel_name,
                CONCAT(u.first_name, ' ', u.last_name) as guest_name
            FROM bookings b
            JOIN hotels h ON b.hotel_id = h.hotel_id
            JOIN users u ON b.user_id = u.user_id
            ORDER BY b.created_at DESC
            LIMIT 5
        ''')
        recent_bookings = cursor.fetchall()
        
        # Get revenue by hotel with currency conversion
        cursor.execute('''
            SELECT 
                h.hotel_name,
                SUM(CASE 
                    WHEN b.currency_id = 1 THEN b.final_amount
                    WHEN b.currency_id = 2 THEN b.final_amount * 0.79
                    WHEN b.currency_id = 3 THEN b.final_amount * 0.85
                    ELSE b.final_amount
                END) as revenue
            FROM bookings b
            JOIN hotels h ON b.hotel_id = h.hotel_id
            WHERE b.status = 'confirmed'
            GROUP BY h.hotel_id, h.hotel_name
            ORDER BY revenue DESC
        ''')
        hotel_revenue = cursor.fetchall()
        
        # Get room occupancy rates
        cursor.execute('''
            SELECT 
                h.hotel_name,
                COUNT(CASE WHEN r.status = 'booked' THEN 1 END) * 100.0 / COUNT(*) as occupancy_rate
            FROM rooms r
            JOIN hotels h ON r.hotel_id = h.hotel_id
            GROUP BY h.hotel_id
        ''')
        occupancy_rates = cursor.fetchall()
        
        # Close cursor after use
        cursor.close()
        
        return render_template('admin/dashboard.html',
                             stats=stats,
                             recent_bookings=recent_bookings,
                             hotel_revenue=hotel_revenue,
                             occupancy_rates=occupancy_rates)
                             
    except Exception as e:
        # Log error
        app.logger.error(f"Hotel list error: {str(e)}\n{traceback.format_exc()}")
        flash('Error loading dashboard', 'error')
        return redirect(url_for('index'))


# Manage bookings route handle bookings management for admin
@app.route('/admin/bookings')
@login_required
@admin_required
def manage_bookings():
    try:
        # Get filter parameters
        status = request.args.get('status', '')
        date_range = request.args.get('date_range', '')
        hotel_id = request.args.get('hotel_id')
        search = request.args.get('search', '')
        page = request.args.get('page', 1, type=int)
        per_page = 2

        cursor = get_db()
        
        # Base query - using table aliases and explicit column selection
        query = '''
            SELECT 
                b.booking_id,
                b.booking_reference,
                b.check_in_date,
                b.check_out_date,
                b.final_amount,
                b.status,
                b.notes,
                b.currency_id,
                b.created_at,
                b.updated_at,
                h.hotel_name,
                CONCAT(u.first_name, ' ', u.last_name) as guest_name,
                u.email as guest_email,
                r.room_number,
                rt.type_name as room_type
            FROM bookings b
            JOIN hotels h ON b.hotel_id = h.hotel_id
            JOIN users u ON b.user_id = u.user_id
            JOIN rooms r ON b.room_id = r.room_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            WHERE 1=1
        '''
        params = []

        # Apply filters
        if status:
            query += ' AND b.status = %s'
            params.append(status)

        # Apply date range filter
        if date_range == 'upcoming':
            query += ' AND b.check_in_date >= CURDATE()'
        elif date_range == 'past':
            query += ' AND b.check_out_date < CURDATE()'
        elif date_range == 'today':
            query += ' AND (DATE(b.check_in_date) = CURDATE() OR DATE(b.check_out_date) = CURDATE())'

        # Apply hotel filter
        if hotel_id:
            query += ' AND b.hotel_id = %s'
            params.append(hotel_id)

        # Apply search filter
        if search:
            query += ''' AND (
                b.booking_reference LIKE %s 
                OR CONCAT(u.first_name, ' ', u.last_name) LIKE %s
                OR u.email LIKE %s
            )'''
            search_term = f'%{search}%'
            params.extend([search_term, search_term, search_term])

        # Get total count for pagination
        count_query = f"SELECT COUNT(*) as total FROM ({query}) as subquery"
        cursor.execute(count_query, params)
        total_bookings = cursor.fetchone()['total']
        total_pages = (total_bookings + per_page - 1) // per_page

        # Add pagination
        query += ' ORDER BY b.created_at DESC LIMIT %s OFFSET %s'
        params.extend([per_page, (page - 1) * per_page])

        # Get bookings
        cursor.execute(query, params)
        bookings = cursor.fetchall()

        # Get hotels for filter
        cursor.execute('SELECT hotel_id, hotel_name FROM hotels ORDER BY hotel_name')
        hotels = cursor.fetchall()

        cursor.execute('SELECT room_type_id, type_name FROM room_types')
        room_types = cursor.fetchall()

        cursor.close()

        return render_template('admin/manage_bookings.html',
                             bookings=bookings,
                             hotels=hotels,
                             room_types=room_types,
                             page=page,
                             total_pages=total_pages,
                             total_bookings=total_bookings,
                             status=status,
                             date_range=date_range,
                             hotel_id=hotel_id,
                             search=search)

    except Exception as e:
        app.logger.error(f'Manage bookings error: {str(e)}')
        flash('Error loading bookings', 'error')
        return redirect(url_for('admin_dashboard'))
    

# Update booking status route handle booking status update for admin
@app.route('/admin/bookings/<int:booking_id>/update-status', methods=['POST'])
@login_required
@admin_required
def update_booking_status(booking_id):
    try:
        new_status = request.form.get('status')
        notes = request.form.get('notes')

        cursor = get_db()
        
        # Update booking status
        cursor.execute('''
            UPDATE bookings 
            SET status = %s, 
                updated_at = CURRENT_TIMESTAMP
            WHERE booking_id = %s
        ''', (new_status, booking_id))

        # Record status change in history with notes
        cursor.execute('''
            INSERT INTO booking_status_history 
            (booking_id, status, changed_by_user_id, notes)
            VALUES (%s, %s, %s, %s)
        ''', (booking_id, new_status, session['user_id'], notes))

        mysql.connection.commit()
        cursor.close()

        flash(f'Booking status updated to {new_status}', 'success')
        return jsonify({'success': True})

    except Exception as e:
        app.logger.error(f'Update booking status error: {str(e)}')
        return jsonify({'success': False, 'message': 'Error updating booking status'}), 500
    

# Edit booking route handle booking edit for admin
@app.route('/admin/bookings/<int:booking_id>/edit', methods=['POST'])
@login_required
@admin_required
def edit_booking(booking_id):
    try:
        # Get form data
        guest_name = request.form.get('guest_name')
        guest_email = request.form.get('guest_email')
        check_in_date = request.form.get('check_in_date')
        check_out_date = request.form.get('check_out_date')
        room_type_id = request.form.get('room_type_id')
        notes = request.form.get('notes')

        cursor = get_db()

        # Update guest information
        cursor.execute('''
            UPDATE booking_details 
            SET guest_name = %s, guest_email = %s 
            WHERE booking_id = %s AND is_primary_guest = TRUE
        ''', (guest_name, guest_email, booking_id))

        # Update booking dates and notes
        cursor.execute('''
            UPDATE bookings 
            SET check_in_date = %s, 
                check_out_date = %s, 
                notes = %s,
                updated_at = CURRENT_TIMESTAMP 
            WHERE booking_id = %s
        ''', (check_in_date, check_out_date, notes, booking_id))

        # Find available room of selected type
        cursor.execute('''
            SELECT r.room_id 
            FROM rooms r 
            WHERE r.room_type_id = %s 
            AND r.status = 'available'
            LIMIT 1
        ''', (room_type_id,))
        
        new_room = cursor.fetchone()
        if new_room:
            cursor.execute('''
                UPDATE bookings 
                SET room_id = %s 
                WHERE booking_id = %s
            ''', (new_room['room_id'], booking_id))

        mysql.connection.commit()
        cursor.close()

        flash('Booking updated successfully', 'success')
        return redirect(url_for('manage_bookings'))

    except Exception as e:
        app.logger.error(f'Edit booking error: {str(e)}')
        flash('Error updating booking', 'error')
        return redirect(url_for('manage_bookings'))
    

# View invoice route handle invoice view for admin with booking details and pricing calculations
@app.route('/admin/bookings/<int:booking_id>/invoice')
@login_required
@admin_required
def view_invoice(booking_id):
    try:
        cursor = get_db()
        
        # Get basic booking details with DATEDIFF calculation
        cursor.execute('''
            SELECT 
                b.*,
                h.hotel_name,
                r.room_number,
                r.base_price,
                rt.type_name as room_type,
                rt.base_price_multiplier,
                bd.guest_name,
                bd.guest_email,
                DATEDIFF(b.check_out_date, b.check_in_date) as total_nights,
                DATEDIFF(b.check_in_date, b.booking_date) as days_in_advance
            FROM bookings b
            JOIN hotels h ON b.hotel_id = h.hotel_id
            JOIN rooms r ON b.room_id = r.room_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            JOIN booking_details bd ON b.booking_id = bd.booking_id
            WHERE b.booking_id = %s AND bd.is_primary_guest = TRUE
        ''', (booking_id,))
        
        booking = cursor.fetchone()
        
        # Check if booking is found
        if not booking:
            flash('Booking not found', 'error')
            return redirect(url_for('manage_bookings'))

        # Calculate pricing
        base_price = float(booking['base_price'])
        base_price_multiplier = float(booking['base_price_multiplier'])
        total_nights = booking['total_nights']
        days_in_advance = booking['days_in_advance']

        # Check if peak season
        check_in_month = booking['check_in_date'].month
        is_peak = check_in_month in [4, 5, 6, 7, 8, 11, 12]
        
        # Calculate base price
        base_price = base_price if is_peak else base_price * 0.5
        base_price *= base_price_multiplier

        # Calculate discount
        if 80 <= days_in_advance <= 90:
            discount_percentage = 30
        elif 60 <= days_in_advance <= 79:
            discount_percentage = 20
        elif 45 <= days_in_advance <= 59:
            discount_percentage = 10
        else:
            discount_percentage = 0

        # Calculate totals
        subtotal = base_price * total_nights
        discount_amount = subtotal * (discount_percentage / 100)
        total_price = subtotal - discount_amount

        # Add calculated values to booking dict
        booking['base_price_per_night'] = base_price
        booking['subtotal'] = subtotal
        booking['discount_percentage'] = discount_percentage
        booking['discount_amount'] = discount_amount
        booking['total_price'] = total_price
        booking['is_peak'] = is_peak

        cursor.close()
        return render_template('admin/invoice.html', booking=booking)

    except Exception as e:
        app.logger.error(f'Invoice error: {str(e)}')
        app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash('Error viewing invoice', 'error')
        return redirect(url_for('manage_bookings'))
    

# Export bookings to Excel route handle bookings export for admin
@app.route('/admin/bookings/export/excel')
@login_required
@admin_required
def export_bookings_excel():
    try:
        # Get filter parameters
        status = request.args.get('status')
        date_range = request.args.get('date_range')
        hotel_id = request.args.get('hotel_id')
        search = request.args.get('search')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')

        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"

        # Define headers
        headers = [
            'Reference',
            'Guest Name',
            'Guest Email',
            'Hotel',
            'Room Number',
            'Room Type',
            'Check-in Date',
            'Check-out Date',
            'Amount',
            'Status',
            'Booking Date',
            'Notes'
        ]

        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Build the SQL query with filters
        query = '''
            SELECT 
                b.booking_reference,
                bd.guest_name,
                bd.guest_email,
                h.hotel_name,
                r.room_number,
                rt.type_name as room_type,
                b.check_in_date,
                b.check_out_date,
                b.final_amount,
                b.status,
                b.booking_date,
                b.notes
            FROM bookings b
            JOIN hotels h ON b.hotel_id = h.hotel_id
            JOIN rooms r ON b.room_id = r.room_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            JOIN booking_details bd ON b.booking_id = bd.booking_id AND bd.is_primary_guest = TRUE
            WHERE 1=1
        '''
        
        params = []

        if status:
            query += " AND b.status = %s"
            params.append(status)
            
        if hotel_id:
            query += " AND b.hotel_id = %s"
            params.append(hotel_id)
            
        if search:
            query += """ AND (
                b.booking_reference LIKE %s OR 
                bd.guest_name LIKE %s OR 
                bd.guest_email LIKE %s
            )"""
            search_term = f"%{search}%"
            params.extend([search_term, search_term, search_term])

        if date_range == 'upcoming':
            query += " AND b.check_in_date > CURDATE()"
        elif date_range == 'today':
            query += " AND DATE(b.check_in_date) = CURDATE()"
        elif date_range == 'past':
            query += " AND b.check_out_date < CURDATE()"
        elif date_range == 'custom' and date_from and date_to:
            query += " AND b.check_in_date BETWEEN %s AND %s"
            params.extend([date_from, date_to])

        query += " ORDER BY b.booking_date DESC"

        # Execute query
        cursor = get_db()
        cursor.execute(query, params)
        bookings = cursor.fetchall()

        # Write data
        for row, booking in enumerate(bookings, 2):
            ws.cell(row=row, column=1).value = booking['booking_reference']
            ws.cell(row=row, column=2).value = booking['guest_name']
            ws.cell(row=row, column=3).value = booking['guest_email']
            ws.cell(row=row, column=4).value = booking['hotel_name']
            ws.cell(row=row, column=5).value = booking['room_number']
            ws.cell(row=row, column=6).value = booking['room_type']
            ws.cell(row=row, column=7).value = booking['check_in_date'].strftime('%Y-%m-%d')
            ws.cell(row=row, column=8).value = booking['check_out_date'].strftime('%Y-%m-%d')
            ws.cell(row=row, column=9).value = float(booking['final_amount'])
            ws.cell(row=row, column=10).value = booking['status']
            ws.cell(row=row, column=11).value = booking['booking_date'].strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row, column=12).value = booking['notes']

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'bookings_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    except Exception as e:
        app.logger.error(f'Excel export error: {str(e)}')
        flash('Error exporting bookings to Excel', 'error')
        return redirect(url_for('manage_bookings'))


# Manage hotels route handle hotels management for admin
@app.route('/admin/hotels')
@login_required
@admin_required
def manage_hotels():
    try:
        cursor = get_db()

        # Pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = 6  # Number of hotels per page
        
        # Get filter parameters
        status = request.args.get('status')
        city = request.args.get('city')
        search = request.args.get('search', '')

        # Base count query
        count_query = '''
            SELECT COUNT(DISTINCT h.hotel_id) as total
            FROM hotels h
            WHERE 1=1
        '''
        count_params = []

        # Base query with room counts
        query = '''
            SELECT 
                h.*,
                COUNT(DISTINCT r.room_id) as current_rooms,
                COUNT(DISTINCT CASE WHEN rt.type_name = 'Standard' THEN r.room_id END) as current_standard_rooms,
                COUNT(DISTINCT CASE WHEN rt.type_name = 'Double' THEN r.room_id END) as current_double_rooms,
                COUNT(DISTINCT CASE WHEN rt.type_name = 'Family' THEN r.room_id END) as current_family_rooms,
                COUNT(DISTINCT CASE WHEN r.status = 'available' THEN r.room_id END) as available_rooms,
                COALESCE(COUNT(DISTINCT CASE WHEN b.status = 'confirmed' 
                    AND b.check_out_date >= CURDATE() THEN b.booking_id END), 0) as active_bookings
            FROM hotels h
            LEFT JOIN rooms r ON h.hotel_id = r.hotel_id
            LEFT JOIN room_types rt ON r.room_type_id = rt.room_type_id
            LEFT JOIN bookings b ON h.hotel_id = b.hotel_id
            WHERE 1=1
        '''
        params = []

        # Apply filters to both queries
        if status:
            query += ' AND h.status = %s'
            count_query += ' AND h.status = %s'
            params.append(status)
            count_params.append(status)

        # Apply city filter
        if city:
            query += ' AND h.city = %s'
            count_query += ' AND h.city = %s'
            params.append(city)
            count_params.append(city)

        # Apply search filter
        if search:
            query += ' AND (h.hotel_name LIKE %s OR h.address LIKE %s)'
            count_query += ' AND (h.hotel_name LIKE %s OR h.address LIKE %s)'
            search_term = f'%{search}%'
            params.extend([search_term, search_term])
            count_params.extend([search_term, search_term])

        # Get total count
        cursor.execute(count_query, count_params)
        total_hotels = cursor.fetchone()['total'] # Fetch total number of hotels
        total_pages = (total_hotels + per_page - 1) // per_page # Calculate total number of pages
            
        # Add grouping and pagination
        query += ' GROUP BY h.hotel_id ORDER BY h.hotel_name LIMIT %s OFFSET %s'
        params.extend([per_page, (page - 1) * per_page]) # Add pagination parameters
        
        # Execute main query
        cursor.execute(query, params)
        hotels = cursor.fetchall() # Fetch all hotels
        
        # Get unique cities for filter
        cursor.execute('SELECT DISTINCT city FROM hotels ORDER BY city')
        cities = cursor.fetchall() # Fetch all unique cities    
        
        # Get amenities for each hotel
        for hotel in hotels:
            cursor.execute('''
                SELECT amenity_name, description, icon_class
                FROM hotel_amenities 
                WHERE hotel_id = %s
            ''', (hotel['hotel_id'],))
            hotel['amenities'] = cursor.fetchall() # Fetch amenities for each hotel
        
        cursor.close()
        
        return render_template('admin/manage_hotels.html',
                             hotels=hotels,
                             cities=cities,
                             current_city=city,
                             current_status=status,
                             search=search,
                             page=page,
                             total_pages=total_pages,
                             total_hotels=total_hotels) # Pass all necessary data to the template
                             
    except Exception as e:
        app.logger.error(f'Manage hotels error: {str(e)}')
        flash('Error loading hotels', 'error')
        return redirect(url_for('admin_dashboard')) # Redirect to admin dashboard if error occurs
    

# Add hotel route handle hotel addition for admin
@app.route('/admin/hotels/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_hotel():
    try:
        cursor = get_db()
        
        # Fetch available amenities
        cursor.execute('''
            SELECT DISTINCT amenity_name, icon_class 
            FROM hotel_amenities 
            ORDER BY amenity_name
        ''')
        amenities = cursor.fetchall()
        
        if request.method == 'POST':
            # Get form data
            hotel_name = request.form.get('hotel_name')
            city = request.form.get('city')
            address = request.form.get('address')
            description = request.form.get('description')
            contact_number = request.form.get('contact_number')
            email = request.form.get('email')
            check_in_time = request.form.get('check_in_time')
            check_out_time = request.form.get('check_out_time')

            # Get capacity data
            total_capacity = int(request.form.get('total_capacity', 0))
            
            # Calculate room distribution
            standard_rooms = round(total_capacity * 0.3)  # 30%
            double_rooms = round(total_capacity * 0.5)    # 50%
            family_rooms = round(total_capacity * 0.2)    # 20%
            
            # Handle image upload
            hotel_image = request.files.get('hotel_image')
            if hotel_image:
                filename = secure_filename(hotel_image.filename)
                hotel_image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                image_path = f'/static/uploads/{filename}'
            else:
                image_path = None
            
            cursor = get_db()
            
            # Insert hotel
            cursor.execute('''
                INSERT INTO hotels (
                    hotel_name, city, address, description,
                    contact_number, email, hotel_image,
                    check_in_time, check_out_time,
                    total_capacity, standard_rooms, double_rooms, family_rooms
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (
                hotel_name, city, address, description,
                contact_number, email, image_path,
                check_in_time, check_out_time,
                total_capacity, standard_rooms, double_rooms, family_rooms
            ))
            
            hotel_id = cursor.lastrowid # Get the last inserted hotel ID
            
            # Add amenities
            amenities = request.form.getlist('amenities[]') # Get selected amenities
            for amenity in amenities:
                cursor.execute('''
                    INSERT INTO hotel_amenities (hotel_id, amenity_name, icon_class)
                    VALUES (%s, %s, %s)
                ''', (hotel_id, amenity, get_amenity_icon(amenity)))
            
            mysql.connection.commit() # Commit changes
            cursor.close() # Close cursor   
            
            flash('Hotel added successfully', 'success') # Show success message
            return redirect(url_for('manage_hotels')) # Redirect to manage hotels page  
            
        return render_template('admin/add_hotel.html', amenities=amenities) # Render add hotel template
        
    except Exception as e:
        app.logger.error(f'Add hotel error: {str(e)}')
        flash('Error adding hotel', 'error')
        return redirect(url_for('manage_hotels'))
    

# Delete hotel route handle hotel deletion for admin
@app.route('/admin/hotels/<int:hotel_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_hotel(hotel_id):
    try:
        cursor = get_db() # Get database cursor
        
        # Check if hotel has any active bookings
        cursor.execute('''
            SELECT COUNT(*) as active_bookings 
            FROM bookings 
            WHERE hotel_id = %s 
            AND status = 'confirmed' 
            AND check_out_date >= CURDATE()
        ''', (hotel_id,))
        
        # Check if hotel has any active bookings
        if cursor.fetchone()['active_bookings'] > 0:
            flash('Cannot delete hotel with active bookings', 'error')
            return redirect(url_for('manage_hotels'))
        
        # Delete records in the correct order to respect foreign key constraints
        
        # 1. Delete room feature mappings first
        cursor.execute('''
            DELETE rfm FROM room_feature_mapping rfm
            INNER JOIN rooms r ON rfm.room_id = r.room_id
            WHERE r.hotel_id = %s
        ''', (hotel_id,))
        
        # 2. Delete bookings
        cursor.execute('DELETE FROM bookings WHERE hotel_id = %s', (hotel_id,))
        
        # 3. Delete rooms
        cursor.execute('DELETE FROM rooms WHERE hotel_id = %s', (hotel_id,))
        
        # 4. Delete hotel amenities
        cursor.execute('DELETE FROM hotel_amenities WHERE hotel_id = %s', (hotel_id,))
        
        # 5. Delete seasonal rates
        cursor.execute('DELETE FROM seasonal_rates WHERE hotel_id = %s', (hotel_id,))
        
        # 6. Delete gallery images
        cursor.execute('DELETE FROM gallery WHERE hotel_id = %s', (hotel_id,))
        
        # 7. Finally delete the hotel
        cursor.execute('DELETE FROM hotels WHERE hotel_id = %s', (hotel_id,))
        
        mysql.connection.commit()
        flash('Hotel deleted successfully', 'success')
        
    except Exception as e:
        app.logger.error(f'Delete hotel error: {str(e)}')
        flash('Error deleting hotel', 'error')
        mysql.connection.rollback()
        
    return redirect(url_for('manage_hotels'))


# Edit hotel route handle hotel editing for admin
@app.route('/admin/hotels/<int:hotel_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_hotel(hotel_id):
    try:
        cursor = get_db() # Get database cursor
        
        if request.method == 'POST':
            # Get form data
            hotel_name = request.form.get('hotel_name')
            city = request.form.get('city')
            address = request.form.get('address')
            title = request.form.get('title')
            description = request.form.get('description')
            contact_number = request.form.get('contact_number')
            email = request.form.get('email')
            check_in_time = request.form.get('check_in_time')
            check_out_time = request.form.get('check_out_time')
            status = request.form.get('status', 'active')
            
            # Get capacity data
            total_capacity = int(request.form.get('total_capacity', 0))
            
            # Calculate room distribution
            standard_rooms = round(total_capacity * 0.3)
            double_rooms = round(total_capacity * 0.5)
            family_rooms = round(total_capacity * 0.2)
            
            # Check if new capacity is less than current room count
            cursor.execute('''
                SELECT COUNT(*) as room_count 
                FROM rooms 
                WHERE hotel_id = %s
            ''', (hotel_id,))
            current_rooms = cursor.fetchone()['room_count']
            
            # Check if current room count is greater than new capacity
            if current_rooms > total_capacity:
                flash('Cannot reduce capacity below current room count', 'error')
                return redirect(url_for('edit_hotel', hotel_id=hotel_id))
            
            # Handle image upload
            if 'hotel_image' in request.files:
                hotel_image = request.files['hotel_image']
                if hotel_image:
                    filename = secure_filename(hotel_image.filename)
                    hotel_image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    image_path = f'/static/uploads/{filename}'
                    
                    cursor.execute('''
                        UPDATE hotels 
                        SET hotel_image = %s 
                        WHERE hotel_id = %s
                    ''', (image_path, hotel_id))
            
            # Update hotel details
            cursor.execute('''
                UPDATE hotels 
                SET hotel_name = %s,
                    city = %s,
                    address = %s,
                    title = %s,
                    description = %s,
                    contact_number = %s,
                    email = %s,
                    check_in_time = %s,
                    check_out_time = %s,
                    status = %s,
                    total_capacity = %s,
                    standard_rooms = %s,
                    double_rooms = %s,
                    family_rooms = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE hotel_id = %s
            ''', (
                hotel_name, city, address, title, description,
                contact_number, email, check_in_time, check_out_time,
                status, total_capacity, standard_rooms, double_rooms,
                family_rooms, hotel_id
            ))

            # Update amenities
            cursor.execute('DELETE FROM hotel_amenities WHERE hotel_id = %s', (hotel_id,))
            amenities = request.form.getlist('amenities[]')
            for amenity in amenities:
                cursor.execute('''
                    INSERT INTO hotel_amenities (hotel_id, amenity_name, icon_class)
                    SELECT %s, amenity_name, icon_class 
                    FROM hotel_amenities 
                    WHERE amenity_name = %s
                ''', (hotel_id, amenity))
            
            mysql.connection.commit()
            flash('Hotel updated successfully', 'success')
            return redirect(url_for('manage_hotels'))
            
        # Get hotel details for editing
        cursor.execute('''
            SELECT h.*,
                   COUNT(r.room_id) as current_rooms,
                   COUNT(CASE WHEN rt.type_name = 'Standard' THEN 1 END) as current_standard_rooms,
                   COUNT(CASE WHEN rt.type_name = 'Double' THEN 1 END) as current_double_rooms,
                   COUNT(CASE WHEN rt.type_name = 'Family' THEN 1 END) as current_family_rooms
            FROM hotels h
            LEFT JOIN rooms r ON h.hotel_id = r.hotel_id
            LEFT JOIN room_types rt ON r.room_type_id = rt.room_type_id
            WHERE h.hotel_id = %s
            GROUP BY h.hotel_id
        ''', (hotel_id,))
        
        hotel = cursor.fetchone()

        # Fetch all available amenities
        cursor.execute('SELECT DISTINCT amenity_name, icon_class FROM hotel_amenities ORDER BY amenity_name')
        all_amenities = cursor.fetchall()
        
        # Fetch current hotel amenities
        cursor.execute('SELECT amenity_name FROM hotel_amenities WHERE hotel_id = %s', (hotel_id,))
        current_amenities = [row['amenity_name'] for row in cursor.fetchall()]

        cursor.close()
        
        if not hotel:
            flash('Hotel not found', 'error')
            return redirect(url_for('manage_hotels'))
            
        return render_template('admin/edit_hotel.html', 
                               hotel=hotel,
                               all_amenities=all_amenities,
                               current_amenities=current_amenities)
        
    except Exception as e:
        app.logger.error(f'Edit hotel error: {str(e)}')
        flash('Error updating hotel', 'error')
        return redirect(url_for('manage_hotels'))
    

# Manage rooms route handle room management for admin
@app.route('/admin/rooms')
@login_required
@admin_required
def manage_rooms():
    try:
        cursor = get_db()
        
        # Pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = 10  # Number of rooms per page
        
        # Get filter parameters
        hotel_id = request.args.get('hotel_id')
        room_type = request.args.get('room_type')
        status = request.args.get('status')
        floor = request.args.get('floor')
        
        # Base query for counting total rooms
        count_query = '''
            SELECT COUNT(*) as total
            FROM rooms r
            JOIN hotels h ON r.hotel_id = h.hotel_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            WHERE 1=1
        '''
        
        # Base query for fetching rooms
        query = '''
            SELECT 
                r.*,
                h.hotel_name,
                rt.type_name,
                COALESCE(b.booking_count, 0) as active_bookings
            FROM rooms r
            JOIN hotels h ON r.hotel_id = h.hotel_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            LEFT JOIN (
                SELECT room_id, COUNT(*) as booking_count
                FROM bookings
                WHERE status = 'confirmed' 
                AND check_out_date >= CURDATE()
                GROUP BY room_id
            ) b ON r.room_id = b.room_id
            WHERE 1=1
        '''
        
        params = []
        
        # Apply filters
        if hotel_id:
            query += ' AND r.hotel_id = %s'
            count_query += ' AND r.hotel_id = %s'
            params.append(hotel_id)
            
        if room_type:
            query += ' AND rt.type_name = %s'
            count_query += ' AND rt.type_name = %s'
            params.append(room_type)
            
        if status:
            query += ' AND r.status = %s'
            count_query += ' AND r.status = %s'
            params.append(status)
            
        if floor:
            query += ' AND r.floor_number = %s'
            count_query += ' AND r.floor_number = %s'
            params.append(floor)
            
        # Add ordering
        query += ' ORDER BY h.hotel_name, r.room_number'
        
        # Add pagination
        query += ' LIMIT %s OFFSET %s'
        pagination_params = params.copy()
        pagination_params.extend([per_page, (page - 1) * per_page])
        
        # Get total count
        cursor.execute(count_query, params)
        total_rooms = cursor.fetchone()['total']
        total_pages = (total_rooms + per_page - 1) // per_page
        
        # Get paginated rooms
        cursor.execute(query, pagination_params)
        rooms = cursor.fetchall()
        
        # Get hotels for filter
        cursor.execute('SELECT hotel_id, hotel_name FROM hotels ORDER BY hotel_name')
        hotels = cursor.fetchall()
        
        # Get room types for filter
        cursor.execute('SELECT type_name FROM room_types ORDER BY type_name')
        room_types = cursor.fetchall()
        
        # Get room features for each room
        for room in rooms:
            cursor.execute('''
                SELECT rf.feature_name, rf.icon_class
                FROM room_feature_mapping rfm
                JOIN room_features rf ON rfm.feature_id = rf.feature_id
                WHERE rfm.room_id = %s
            ''', (room['room_id'],))
            room['features'] = cursor.fetchall()
        
        cursor.close()
        
        return render_template('admin/manage_rooms.html',
                             rooms=rooms,
                             hotels=hotels,
                             room_types=room_types,
                             current_hotel=hotel_id,
                             current_type=room_type,
                             current_status=status,
                             current_floor=floor,
                             page=page,
                             total_pages=total_pages,
                             total_rooms=total_rooms)
                             
    except Exception as e:
        app.logger.error(f'Manage rooms error: {str(e)}')
        flash('Error loading rooms', 'error')
        return redirect(url_for('admin_dashboard'))


# Add room route handle room addition for admin
@app.route('/admin/rooms/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_room():
    try:
        cursor = get_db()
        
        if request.method == 'POST':
            # Get form data
            hotel_id = request.form.get('hotel_id')
            room_type_id = request.form.get('room_type_id')
            room_number = request.form.get('room_number')
            floor_number = request.form.get('floor_number')
            base_price = request.form.get('base_price')
            
            # Handle image upload
            room_image = request.files.get('room_image')
            if room_image:
                filename = secure_filename(room_image.filename)
                room_image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                image_path = f'/static/uploads/{filename}'
            else:
                image_path = None
            
            # Insert room
            cursor.execute('''
                INSERT INTO rooms (
                    hotel_id, room_type_id, room_number, floor_number,
                    base_price, room_image, status
                ) VALUES (%s, %s, %s, %s, %s, %s, 'available')
            ''', (hotel_id, room_type_id, room_number, floor_number, 
                  base_price, image_path))
            
            room_id = cursor.lastrowid
            
            # Add features
            features = request.form.getlist('features[]')
            for feature_id in features:
                cursor.execute('''
                    INSERT INTO room_feature_mapping (room_id, feature_id)
                    VALUES (%s, %s)
                ''', (room_id, feature_id))
            
            mysql.connection.commit()
            flash('Room added successfully', 'success')
            return redirect(url_for('manage_rooms'))
        
        # Get hotels and room types for form
        cursor.execute('SELECT hotel_id, hotel_name FROM hotels ORDER BY hotel_name')
        hotels = cursor.fetchall() # Fetch all hotels  
        
        cursor.execute('SELECT * FROM room_types ORDER BY type_name')
        room_types = cursor.fetchall() # Fetch all room types
        
        cursor.execute('SELECT * FROM room_features ORDER BY feature_name')
        features = cursor.fetchall() # Fetch all room features
        
        cursor.close()
        
        return render_template('admin/add_room.html',
                             hotels=hotels,
                             room_types=room_types,
                             features=features)
                             
    except Exception as e:
        app.logger.error(f'Add room error: {str(e)}')
        flash('Error adding room', 'error')
        return redirect(url_for('manage_rooms'))
    

# Delete room route handle room deletion for admin
@app.route('/admin/rooms/<int:room_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_room(room_id):
    try:
        cursor = get_db()
        
        # Check if room has any active bookings
        cursor.execute('''
            SELECT COUNT(*) as active_bookings 
            FROM bookings 
            WHERE room_id = %s 
            AND status = 'confirmed' 
            AND check_out_date >= CURDATE()
        ''', (room_id,))
        
        if cursor.fetchone()['active_bookings'] > 0: # Check if room has any active bookings
            flash('Cannot delete room with active bookings', 'error')
            return redirect(url_for('manage_rooms'))
        
        # Delete records in the correct order to respect foreign key constraints
        
        # 1. Delete room feature mappings
        cursor.execute('DELETE FROM room_feature_mapping WHERE room_id = %s', (room_id,))
        
        # 2. Delete bookings
        cursor.execute('DELETE FROM bookings WHERE room_id = %s', (room_id,))
        
        # 3. Finally delete the room
        cursor.execute('DELETE FROM rooms WHERE room_id = %s', (room_id,))
        
        mysql.connection.commit()
        cursor.close()
        flash('Room deleted successfully', 'success')
        
    except Exception as e:
        app.logger.error(f'Delete room error: {str(e)}')
        flash('Error deleting room', 'error')
        mysql.connection.rollback()
        
    return redirect(url_for('manage_rooms'))


# Edit room route handle room editing for admin
@app.route('/admin/rooms/<int:room_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_room(room_id):
    try:
        cursor = get_db()
        
        # Handle POST request for updating room details
        if request.method == 'POST':
            # Update room details
            cursor.execute('''
                UPDATE rooms 
                SET hotel_id = %s,
                    room_type_id = %s,
                    room_number = %s,
                    floor_number = %s,
                    base_price = %s,
                    status = %s,
                    is_active = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE room_id = %s
            ''', (
                request.form.get('hotel_id'),
                request.form.get('room_type_id'),
                request.form.get('room_number'),
                request.form.get('floor_number'),
                request.form.get('base_price'),
                request.form.get('status'),
                request.form.get('is_active', 'off') == 'on',
                room_id
            ))
            
            # Handle image update if provided
            if 'room_image' in request.files:
                room_image = request.files['room_image']
                if room_image:
                    filename = secure_filename(room_image.filename)
                    room_image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    image_path = f'/static/uploads/{filename}'
                    cursor.execute('UPDATE rooms SET room_image = %s WHERE room_id = %s',
                                 (image_path, room_id))
            
            # Update features
            cursor.execute('DELETE FROM room_feature_mapping WHERE room_id = %s', 
                         (room_id,))
            features = request.form.getlist('features[]')
            for feature_id in features:
                cursor.execute('''
                    INSERT INTO room_feature_mapping (room_id, feature_id)
                    VALUES (%s, %s)
                ''', (room_id, feature_id))
            
            mysql.connection.commit()
            flash('Room updated successfully', 'success')
            return redirect(url_for('manage_rooms'))
        
        # Get room details for editing
        cursor.execute('''
            SELECT r.*, h.hotel_name, rt.type_name
            FROM rooms r
            JOIN hotels h ON r.hotel_id = h.hotel_id
            JOIN room_types rt ON r.room_type_id = rt.room_type_id
            WHERE r.room_id = %s
        ''', (room_id,))
        room = cursor.fetchone()
        
        if not room:
            flash('Room not found', 'error')
            return redirect(url_for('manage_rooms'))
        
        # Get selected features
        cursor.execute('''
            SELECT feature_id 
            FROM room_feature_mapping 
            WHERE room_id = %s
        ''', (room_id,))
        selected_features = [row['feature_id'] for row in cursor.fetchall()]
        room['features'] = selected_features
        
        # Get hotels, room types and features for form
        cursor.execute('SELECT hotel_id, hotel_name FROM hotels ORDER BY hotel_name')
        hotels = cursor.fetchall()
        
        cursor.execute('SELECT * FROM room_types ORDER BY type_name')
        room_types = cursor.fetchall()
        
        cursor.execute('SELECT * FROM room_features ORDER BY feature_name')
        features = cursor.fetchall()
        
        cursor.close()
        
        return render_template('admin/edit_room.html',
                             room=room,
                             hotels=hotels,
                             room_types=room_types,
                             features=features)
                             
    except Exception as e:
        app.logger.error(f'Edit room error: {str(e)}')
        flash('Error updating room', 'error')
        return redirect(url_for('manage_rooms'))


# Manage users route handle user management for admin
@app.route('/admin/users')
@login_required
@admin_required
def manage_users():
    try:
        cursor = get_db()
        
        # Pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = 20
        
        # Get filter parameters
        role = request.args.get('role')
        status = request.args.get('status')
        search = request.args.get('search', '')
        
        # Base query
        query = '''
            SELECT 
                u.user_id,
                u.first_name,
                u.last_name,
                u.email,
                u.phone_number,
                u.profile_image,
                u.is_active,
                u.last_login,
                u.created_at,
                u.address,
                ur.role_name,
                COUNT(DISTINCT b.booking_id) as total_bookings,
                MAX(b.booking_date) as last_booking
            FROM users u
            LEFT JOIN user_roles ur ON u.role_id = ur.role_id
            LEFT JOIN bookings b ON u.user_id = b.user_id
            WHERE 1=1
        '''
        params = []
        
        # Apply filters
        if role and role != '':
            query += ' AND ur.role_name = %s'
            params.append(role)
            
        if status and status != '':
            is_active = status == 'active'
            query += ' AND u.is_active = %s'
            params.append(is_active)
            
        if search and search != '' and search.lower() != 'none':
            query += ''' AND (
                u.email LIKE %s 
                OR u.first_name LIKE %s
                OR u.last_name LIKE %s
                OR u.phone_number LIKE %s
            )'''
            search_term = f'%{search}%'
            params.extend([search_term] * 4)
        
        # Create count query
        count_query = f'''
            SELECT COUNT(*) as total FROM (
                SELECT DISTINCT u.user_id
                FROM users u
                LEFT JOIN user_roles ur ON u.role_id = ur.role_id
                LEFT JOIN bookings b ON u.user_id = b.user_id
                WHERE 1=1
        '''
        
        if role and role != '':
            count_query += ' AND ur.role_name = %s'
        if status and status != '':
            count_query += ' AND u.is_active = %s'
        if search and search != '' and search.lower() != 'none':
            count_query += ''' AND (
                u.email LIKE %s 
                OR u.first_name LIKE %s
                OR u.last_name LIKE %s
                OR u.phone_number LIKE %s
            )'''
            
        count_query += ') as subquery'
        
        # Get total count
        cursor.execute(count_query, params)
        total_users = cursor.fetchone()['total']
        total_pages = (total_users + per_page - 1) // per_page
            
        # Add GROUP BY and pagination
        query += '''
            GROUP BY 
                u.user_id,
                u.first_name,
                u.last_name,
                u.email,
                u.phone_number,
                u.profile_image,
                u.is_active,
                u.last_login,
                u.created_at,
                u.address,
                ur.role_name
            ORDER BY u.created_at DESC
            LIMIT %s OFFSET %s
        '''
        params.extend([per_page, (page - 1) * per_page])
        
        cursor.execute(query, params)
        users = cursor.fetchall()
        
        # Get roles for filter
        cursor.execute('SELECT DISTINCT role_name FROM user_roles ORDER BY role_name')
        roles = cursor.fetchall()
        
        cursor.close()
        
        # Handle any None values
        if users:
            for user in users:
                user['first_name'] = user['first_name'] or ''
                user['last_name'] = user['last_name'] or ''
                user['phone_number'] = user['phone_number'] or ''
                user['profile_image'] = user['profile_image'] or ''
                user['address'] = user['address'] or ''
                user['role_name'] = user['role_name'] or 'User'
                user['total_bookings'] = user['total_bookings'] or 0
        
        # Don't pass 'None' as a search term
        search = None if search and search.lower() == 'none' else search
        
        return render_template('admin/manage_users.html',
                             users=users or [], 
                             roles=roles or [], 
                             current_role=role,
                             current_status=status,
                             search=search,
                             page=page,
                             total_pages=total_pages,
                             total_users=total_users)
                             
    except Exception as e:
        app.logger.error(f'Manage users error: {str(e)}')
        flash('Error loading users', 'error')
        return redirect(url_for('admin_dashboard'))
    

# Add user route handle user addition for admin
@app.route('/admin/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    try:
        if request.method == 'POST':
            # Get form inputs
            password = request.form.get('password')
            confirm_password = request.form.get('confirm_password')

            # Validate passwords
            if password != confirm_password:
                flash('Passwords do not match!', 'danger')
                return redirect(url_for('add_user'))

            if not is_password_strong(password):
                flash('Password must be at least 8 characters long, include at least one uppercase letter, one lowercase letter, one number, and one special character.', 'danger')
                return redirect(url_for('add_user'))

            # Hash the password
            password_hash = generate_password_hash(password)

            # Insert new user into database
            cursor = get_db()
            cursor.execute('''
                INSERT INTO users (
                    role_id, first_name, last_name, email, password_hash,
                    phone_number, address, is_active
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ''', (
                request.form.get('role_id'),
                request.form.get('first_name'),
                request.form.get('last_name'),
                request.form.get('email'),
                password_hash,
                request.form.get('phone_number'),
                request.form.get('address'),
                request.form.get('is_active', 'off') == 'on'
            ))

            # Handle profile image if provided
            if 'profile_image' in request.files:
                profile_image = request.files['profile_image']
                if profile_image and profile_image.filename:
                    filename = secure_filename(profile_image.filename)
                    profile_image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    image_path = f'/static/uploads/{filename}'
                    cursor.execute('UPDATE users SET profile_image = %s WHERE user_id = LAST_INSERT_ID()',
                                   (image_path,))

            # Commit the transaction
            mysql.connection.commit()
            cursor.close()

            flash('User added successfully', 'success')
            return redirect(url_for('manage_users'))

        # Get roles for the form
        cursor = get_db()
        cursor.execute('SELECT * FROM user_roles ORDER BY role_name')
        roles = cursor.fetchall()
        cursor.close()

        return render_template('admin/add_user.html', roles=roles)

    except Exception as e:
        app.logger.error(f'Add user error: {str(e)}')
        flash('Error adding user', 'error')
        return redirect(url_for('manage_users'))
    

# Delete user route handle user deletion for admin
@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    try:
        cursor = get_db()
        
        # Get current user's ID from the session
        current_user_id = session.get('user_id')
        
        # Prevent deleting your own account
        if user_id == current_user_id:
            flash('You cannot delete your own account', 'error')
            return redirect(url_for('manage_users'))
            
        # Check if user is an admin
        cursor.execute('''
            SELECT u.user_id, ur.role_name 
            FROM users u
            JOIN user_roles ur ON u.role_id = ur.role_id
            WHERE u.user_id = %s
        ''', (user_id,))
        user = cursor.fetchone()
        
        if user and user['role_name'] == 'admin':
            flash('Admin accounts cannot be deleted', 'error')
            return redirect(url_for('manage_users'))
        
        # Delete records in the correct order to respect foreign key constraints
        
        # 1. Delete audit logs first
        cursor.execute('DELETE FROM audit_logs WHERE user_id = %s', (user_id,))
        
        # 2. Delete user sessions
        cursor.execute('DELETE FROM user_sessions WHERE user_id = %s', (user_id,))
        
        # 3. Delete password reset tokens
        cursor.execute('DELETE FROM password_reset_tokens WHERE user_id = %s', (user_id,))
        
        # 4. Delete email verification tokens
        cursor.execute('DELETE FROM email_verification WHERE user_id = %s', (user_id,))
        
        # 5. Delete notifications
        cursor.execute('DELETE FROM notifications WHERE user_id = %s', (user_id,))
        
        # 6. Delete booking status history entries
        cursor.execute('DELETE FROM booking_status_history WHERE changed_by_user_id = %s', (user_id,))
        
        # 7. Delete booking transactions
        cursor.execute('''
            DELETE bt FROM booking_transactions bt
            INNER JOIN bookings b ON bt.booking_id = b.booking_id
            WHERE b.user_id = %s
        ''', (user_id,))
        
        # 8. Delete booking details
        cursor.execute('''
            DELETE bd FROM booking_details bd
            INNER JOIN bookings b ON bd.booking_id = b.booking_id
            WHERE b.user_id = %s
        ''', (user_id,))
        
        # 9. Delete bookings
        cursor.execute('DELETE FROM bookings WHERE user_id = %s', (user_id,))
        
        # 10. Finally delete the user
        cursor.execute('DELETE FROM users WHERE user_id = %s', (user_id,))
        
        mysql.connection.commit()
        cursor.close()
        flash('User deleted successfully', 'success')
        
    except Exception as e:
        app.logger.error(f'Delete user error: {str(e)}')
        flash('Error deleting user', 'error')
        mysql.connection.rollback()
        
    return redirect(url_for('manage_users'))
    

# Edit user route handle user editing for admin
@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    try:
        cursor = get_db()
        
        if request.method == 'POST':
            # Fetch the existing password hash from the database
            cursor.execute('SELECT password_hash FROM users WHERE user_id = %s', (user_id,))
            existing_user = cursor.fetchone()
            if not existing_user:
                flash('User not found', 'error')
                return redirect(url_for('manage_users'))
            
            existing_password_hash = existing_user['password_hash']
            
            # Update user details
            update_query = '''
                UPDATE users 
                SET role_id = %s,
                    first_name = %s,
                    last_name = %s,
                    email = %s,
                    phone_number = %s,
                    address = %s,
                    is_active = %s,
                    updated_at = CURRENT_TIMESTAMP
            '''
            params = [
                request.form.get('role_id'),
                request.form.get('first_name'),
                request.form.get('last_name'),
                request.form.get('email'),
                request.form.get('phone_number'),
                request.form.get('address'),
                request.form.get('is_active', 'off') == 'on'
            ]
            
            # Update password if provided
            new_password = request.form.get('new_password')
            if new_password:
                # Validate strong password
                if not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&#])[A-Za-z\d@$!%*?&#]{8,}$', new_password):
                    flash('Password must be at least 8 characters long and include an uppercase letter, a lowercase letter, a number, and a special character.', 'error')
                    return redirect(request.url)
                
                # Check if the new password is the same as the existing password
                if check_password_hash(existing_password_hash, new_password):
                    flash('The new password cannot be the same as the current password.', 'error')
                    return redirect(request.url)

                # Hash and update password
                update_query += ', password_hash = %s'
                params.append(generate_password_hash(new_password))
            
            update_query += ' WHERE user_id = %s'
            params.append(user_id)
            
            cursor.execute(update_query, params)
            
            # Handle profile image update if provided
            if 'profile_image' in request.files:
                profile_image = request.files['profile_image']
                if profile_image and profile_image.filename:
                    filename = secure_filename(profile_image.filename)
                    profile_image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    image_path = f'/static/uploads/{filename}'
                    cursor.execute('UPDATE users SET profile_image = %s WHERE user_id = %s',
                                   (image_path, user_id))
            
            mysql.connection.commit()
            flash('User updated successfully', 'success')
            return redirect(url_for('manage_users'))
        
        # Get user details for editing
        cursor.execute('''
            SELECT u.*, ur.role_name
            FROM users u
            JOIN user_roles ur ON u.role_id = ur.role_id
            WHERE u.user_id = %s
        ''', (user_id,))
        user = cursor.fetchone()
        
        if not user:
            flash('User not found', 'error')
            return redirect(url_for('manage_users'))
        
        # Get roles for form
        cursor.execute('SELECT * FROM user_roles ORDER BY role_name')
        roles = cursor.fetchall()
        
        cursor.close()
        return render_template('admin/edit_user.html',
                               user=user,
                               roles=roles)
                             
    except Exception as e:
        app.logger.error(f'Edit user error: {str(e)}')
        flash('Error updating user', 'error')
        return redirect(url_for('manage_users'))


# Reports route handle reports generation for admin
@app.route('/admin/reports')
@login_required
@admin_required
def admin_reports():
    try:
        cursor = get_db()
        
        # Date range filter
        date_range = request.args.get('date_range', 'last_30_days')
        custom_start = request.args.get('start_date')
        custom_end = request.args.get('end_date')
        
        # Set date range based on filter
        if date_range == 'last_7_days':
            start_date = 'DATE_SUB(CURDATE(), INTERVAL 7 DAY)'
        elif date_range == 'last_30_days':
            start_date = 'DATE_SUB(CURDATE(), INTERVAL 30 DAY)'
        elif date_range == 'last_90_days':
            start_date = 'DATE_SUB(CURDATE(), INTERVAL 90 DAY)'
        elif date_range == 'custom' and custom_start and custom_end:
            start_date = f"'{custom_start}'"
            end_date = f"'{custom_end}'"
        else:
            start_date = 'DATE_SUB(CURDATE(), INTERVAL 30 DAY)'
        
        # 1. Booking Statistics
        cursor.execute(f'''
            SELECT 
                COUNT(*) as total_bookings,
                SUM(CASE WHEN status = 'confirmed' THEN 1 ELSE 0 END) as confirmed_bookings,
                SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) as cancelled_bookings,
                COALESCE(AVG(CASE WHEN status = 'confirmed' THEN DATEDIFF(check_out_date, check_in_date) END), 0) as avg_stay_duration,
                COALESCE(SUM(CASE WHEN status = 'confirmed' THEN final_amount ELSE 0 END), 0) as total_revenue,
                COALESCE(AVG(CASE WHEN status = 'confirmed' THEN final_amount END), 0) as avg_booking_value
            FROM bookings
            WHERE booking_date >= {start_date}
        ''')
        booking_stats = cursor.fetchone()
        
        # 2. Revenue by Hotel
        cursor.execute(f'''
            SELECT 
                h.hotel_name,
                COUNT(b.booking_id) as booking_count,
                COALESCE(SUM(CASE WHEN b.status = 'confirmed' THEN b.final_amount ELSE 0 END), 0) as revenue,
                COALESCE(AVG(CASE WHEN b.status = 'confirmed' THEN b.final_amount END), 0) as avg_booking_value
            FROM hotels h
            LEFT JOIN bookings b ON h.hotel_id = b.hotel_id
            WHERE b.booking_date >= {start_date} OR b.booking_date IS NULL
            GROUP BY h.hotel_id
            ORDER BY revenue DESC
        ''')
        hotel_revenue = cursor.fetchall()
        
        # 3. Room Type Performance
        cursor.execute(f'''
            SELECT 
                rt.type_name,
                COUNT(b.booking_id) as booking_count,
                COALESCE(SUM(CASE WHEN b.status = 'confirmed' THEN b.final_amount ELSE 0 END), 0) as revenue,
                COALESCE(AVG(CASE WHEN b.status = 'confirmed' THEN b.final_amount END), 0) as avg_rate
            FROM room_types rt
            LEFT JOIN rooms r ON rt.room_type_id = r.room_type_id
            LEFT JOIN bookings b ON r.room_id = b.room_id
            WHERE b.booking_date >= {start_date} OR b.booking_date IS NULL
            GROUP BY rt.room_type_id
            ORDER BY revenue DESC
        ''')
        room_type_stats = cursor.fetchall()
        
        # 4. Daily Revenue Trend
        cursor.execute(f'''
            SELECT 
                DATE(booking_date) as date,
                COUNT(*) as bookings,
                SUM(CASE WHEN status = 'confirmed' THEN final_amount ELSE 0 END) as revenue
            FROM bookings
            WHERE booking_date >= {start_date}
            GROUP BY DATE(booking_date)
            ORDER BY date
        ''')
        daily_revenue = cursor.fetchall()
        
        # 5. Occupancy Rates
        cursor.execute(f'''
            SELECT 
                h.hotel_name,
                COUNT(DISTINCT r.room_id) as total_rooms,
                COUNT(DISTINCT CASE 
                    WHEN b.status = 'confirmed' 
                    AND CURDATE() BETWEEN b.check_in_date AND b.check_out_date 
                    THEN r.room_id 
                END) as occupied_rooms
            FROM hotels h
            LEFT JOIN rooms r ON h.hotel_id = r.hotel_id
            LEFT JOIN bookings b ON r.room_id = b.room_id
            GROUP BY h.hotel_id
        ''')
        occupancy_rates = cursor.fetchall()
        
        # 6. Top Customers - Updated query
        cursor.execute(f'''
            SELECT 
                CONCAT(u.first_name, ' ', u.last_name) as full_name,
                COUNT(b.booking_id) as booking_count,
                SUM(CASE WHEN b.status = 'confirmed' THEN b.final_amount ELSE 0 END) as total_spent,
                MAX(b.booking_date) as last_booking
            FROM users u
            JOIN bookings b ON u.user_id = b.user_id
            WHERE b.booking_date >= {start_date}
            GROUP BY u.user_id, u.first_name, u.last_name
            ORDER BY total_spent DESC
            LIMIT 10
        ''')
        top_customers = cursor.fetchall()
        
        cursor.close()
        
        return render_template('admin/reports.html',
                             date_range=date_range,
                             custom_start=custom_start,
                             custom_end=custom_end,
                             booking_stats=booking_stats,
                             hotel_revenue=hotel_revenue,
                             room_type_stats=room_type_stats,
                             daily_revenue=daily_revenue,
                             occupancy_rates=occupancy_rates,
                             top_customers=top_customers)
                             
    except Exception as e:
        app.logger.error(f'Reports error: {str(e)}')
        flash('Error generating reports', 'error')
        return redirect(url_for('admin_dashboard'))
    

# Export report PDF route handle PDF export for admin
@app.route('/admin/reports/export/pdf')
@login_required
@admin_required
def export_report_pdf():
    try:
        # Get the same data as in admin_reports
        report_data = get_report_data(
            request.args.get('date_range'),
            request.args.get('start_date'),
            request.args.get('end_date')
        )
        
        # Render the PDF template
        html = render_template(
            'admin/report_template.html',
            **report_data,
            generated_at=datetime.now()
        )
        
        # Generate PDF
        config = pdfkit.configuration(wkhtmltopdf=get_wkhtmltopdf_path())
        options = {
            'page-size': 'A4',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': "UTF-8",
            'no-outline': None,
            'enable-local-file-access': True
        }
        
        pdf = pdfkit.from_string(html, False, options=options, configuration=config)
        
        # Create response
        response = make_response(pdf) # Create response object 
        response.headers['Content-Type'] = 'application/pdf' # Set content type to PDF
        response.headers['Content-Disposition'] = f'attachment; filename=hotel_report_{datetime.now().strftime("%Y%m%d")}.pdf' # Set content disposition to attachment
        
        return response
        
    except Exception as e:
        app.logger.error(f'PDF report generation error: {str(e)}')
        flash('Error generating PDF report', 'error')
        return redirect(url_for('admin_reports'))
    

# Export report Excel route handle Excel export for admin
@app.route('/admin/reports/export/excel')
@login_required
@admin_required
def export_report_excel():
    try:
        # Get report data
        report_data = get_report_data(
            request.args.get('date_range'),
            request.args.get('start_date'),
            request.args.get('end_date')
        )
        
        # Create workbook
        wb = Workbook()
        
        # Booking Statistics
        ws = wb.active
        ws.title = "Booking Statistics"
        headers = ['Metric', 'Value']
        ws.append(headers)
        
        stats = [
            ['Total Bookings', report_data['booking_stats']['total_bookings']],
            ['Confirmed Bookings', report_data['booking_stats']['confirmed_bookings']],
            ['Cancelled Bookings', report_data['booking_stats']['cancelled_bookings']],
            ['Total Revenue', f"£{report_data['booking_stats']['total_revenue']:.2f}"],
            ['Average Booking Value', f"£{report_data['booking_stats']['avg_booking_value']:.2f}"],
            ['Average Stay Duration', f"{report_data['booking_stats']['avg_stay_duration']:.1f} nights"]
        ]
        
        # Append stats to worksheet
        for row in stats:
            ws.append(row)
            
        # Style the worksheet
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        # Hotel Revenue
        ws = wb.create_sheet("Hotel Revenue")
        headers = ['Hotel', 'Bookings', 'Revenue', 'Avg Booking Value']
        ws.append(headers)
        
        for hotel in report_data['hotel_revenue']:
            ws.append([
                hotel['hotel_name'],
                hotel['booking_count'],
                f"£{hotel['revenue']:.2f}",
                f"£{hotel['avg_booking_value']:.2f}" if hotel['avg_booking_value'] else '£0.00'
            ])
            
        # Style the worksheet
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        # Room Type Performance
        ws = wb.create_sheet("Room Performance")
        headers = ['Room Type', 'Bookings', 'Revenue', 'Average Rate']
        ws.append(headers)
        
        # Append room type stats to worksheet
        for room in report_data['room_type_stats']:
            ws.append([
                room['type_name'],
                room['booking_count'],
                f"£{room['revenue']:.2f}",
                f"£{room['avg_rate']:.2f}" if room['avg_rate'] else '£0.00'
            ])
            
        # Style the worksheet
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Send file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'hotel_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        app.logger.error(f'Excel report generation error: {str(e)}')
        flash('Error generating Excel report', 'error')
        return redirect(url_for('admin_reports'))


# Helper function to get report data for both PDF and Excel exports
def get_report_data(date_range, custom_start=None, custom_end=None):
    cursor = get_db()
    
    # Set date range based on filter
    if date_range == 'last_7_days':
        start_date = 'DATE_SUB(CURDATE(), INTERVAL 7 DAY)'
    elif date_range == 'last_30_days':
        start_date = 'DATE_SUB(CURDATE(), INTERVAL 30 DAY)'
    elif date_range == 'last_90_days':
        start_date = 'DATE_SUB(CURDATE(), INTERVAL 90 DAY)'
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = f"'{custom_start}'"
        end_date = f"'{custom_end}'"
    else:
        start_date = 'DATE_SUB(CURDATE(), INTERVAL 30 DAY)'
    
    # 1. Booking Statistics
    cursor.execute(f'''
        SELECT 
            COUNT(*) as total_bookings,
            SUM(CASE WHEN status = 'confirmed' THEN 1 ELSE 0 END) as confirmed_bookings,
            SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) as cancelled_bookings,
            COALESCE(AVG(CASE WHEN status = 'confirmed' THEN DATEDIFF(check_out_date, check_in_date) END), 0) as avg_stay_duration,
            COALESCE(SUM(CASE WHEN status = 'confirmed' THEN final_amount ELSE 0 END), 0) as total_revenue,
            COALESCE(AVG(CASE WHEN status = 'confirmed' THEN final_amount END), 0) as avg_booking_value
        FROM bookings
        WHERE booking_date >= {start_date}
    ''')
    booking_stats = cursor.fetchone()
    
    # 2. Revenue by Hotel
    cursor.execute(f'''
        SELECT 
            h.hotel_name,
            COUNT(b.booking_id) as booking_count,
            COALESCE(SUM(CASE WHEN b.status = 'confirmed' THEN b.final_amount ELSE 0 END), 0) as revenue,
            COALESCE(AVG(CASE WHEN b.status = 'confirmed' THEN b.final_amount END), 0) as avg_booking_value
        FROM hotels h
        LEFT JOIN bookings b ON h.hotel_id = b.hotel_id
        WHERE b.booking_date >= {start_date} OR b.booking_date IS NULL
        GROUP BY h.hotel_id
        ORDER BY revenue DESC
    ''')
    hotel_revenue = cursor.fetchall()
    
    # 3. Room Type Performance
    cursor.execute(f'''
        SELECT 
            rt.type_name,
            COUNT(b.booking_id) as booking_count,
            COALESCE(SUM(CASE WHEN b.status = 'confirmed' THEN b.final_amount ELSE 0 END), 0) as revenue,
            COALESCE(AVG(CASE WHEN b.status = 'confirmed' THEN b.final_amount END), 0) as avg_rate
        FROM room_types rt
        LEFT JOIN rooms r ON rt.room_type_id = r.room_type_id
        LEFT JOIN bookings b ON r.room_id = b.room_id
        WHERE b.booking_date >= {start_date} OR b.booking_date IS NULL
        GROUP BY rt.room_type_id
        ORDER BY revenue DESC
    ''')
    room_type_stats = cursor.fetchall()
    
    # 4. Daily Revenue Trend
    cursor.execute(f'''
        SELECT 
            DATE(booking_date) as date,
            COUNT(*) as bookings,
            SUM(CASE WHEN status = 'confirmed' THEN final_amount ELSE 0 END) as revenue
        FROM bookings
        WHERE booking_date >= {start_date}
        GROUP BY DATE(booking_date)
        ORDER BY date
    ''')
    daily_revenue = cursor.fetchall()
    
    # 5. Occupancy Rates
    cursor.execute(f'''
        SELECT 
            h.hotel_name,
            COUNT(DISTINCT r.room_id) as total_rooms,
            COUNT(DISTINCT CASE 
                WHEN b.status = 'confirmed' 
                AND CURDATE() BETWEEN b.check_in_date AND b.check_out_date 
                THEN r.room_id 
            END) as occupied_rooms
        FROM hotels h
        LEFT JOIN rooms r ON h.hotel_id = r.hotel_id
        LEFT JOIN bookings b ON r.room_id = b.room_id
        GROUP BY h.hotel_id
    ''')
    occupancy_rates = cursor.fetchall()
    
    # 6. Top Customers
    cursor.execute(f'''
        SELECT 
            CONCAT(u.first_name, ' ', u.last_name) as full_name,
            COUNT(b.booking_id) as booking_count,
            SUM(CASE WHEN b.status = 'confirmed' THEN b.final_amount ELSE 0 END) as total_spent,
            MAX(b.booking_date) as last_booking
        FROM users u
        JOIN bookings b ON u.user_id = b.user_id
        WHERE b.booking_date >= {start_date}
        GROUP BY u.user_id, u.first_name, u.last_name
        ORDER BY total_spent DESC
        LIMIT 10
    ''')
    top_customers = cursor.fetchall()
    
    cursor.close()
    
    return {
        'booking_stats': booking_stats,
        'hotel_revenue': hotel_revenue,
        'room_type_stats': room_type_stats,
        'daily_revenue': daily_revenue,
        'occupancy_rates': occupancy_rates,
        'top_customers': top_customers
    }


# Manage currencies route handle currency management for admin
@app.route('/admin/currencies', methods=['GET', 'POST'])
@admin_required
def admin_currencies():
    try:
        cursor = get_db() # Get database cursor
        
        if request.method == 'POST':
            action = request.form.get('action') # Get action from form
            
            if action == 'toggle_currency':
                currency_id = request.form.get('currency_id') # Get currency ID from form
                cursor.execute('UPDATE currencies SET is_active = NOT is_active WHERE currency_id = %s', (currency_id,)) # Toggle currency status
                mysql.connection.commit() # Commit changes
                flash('Currency status updated successfully', 'success') # Show success message
                return redirect(url_for('admin_currencies')) # Redirect to currencies page
                
            elif action == 'update_rate':
                from_currency = request.form.get('from_currency') # Get from currency from form
                to_currency = request.form.get('to_currency') # Get to currency from form
                rate = float(request.form.get('rate')) # Get rate from form
                
                # Update exchange rate for direct conversion
                cursor.execute('''
                    UPDATE exchange_rates 
                    SET rate = %s, last_updated = CURRENT_TIMESTAMP 
                    WHERE from_currency_id = %s AND to_currency_id = %s
                ''', (rate, from_currency, to_currency))
                
                # Update exchange rate for reverse conversion
                cursor.execute('''
                    UPDATE exchange_rates 
                    SET rate = %s, last_updated = CURRENT_TIMESTAMP 
                    WHERE from_currency_id = %s AND to_currency_id = %s
                ''', (1/rate, to_currency, from_currency))
                mysql.connection.commit()
                flash('Exchange rate updated successfully', 'success')
                return redirect(url_for('admin_currencies')) # Redirect to currencies page
            
            elif action == 'add_currency':
                code = request.form.get('code').upper() # Get code from form
                name = request.form.get('name') # Get name from form
                symbol = request.form.get('symbol') # Get symbol from form
                
                # Check if currency code already exists
                cursor.execute('SELECT COUNT(*) as count FROM currencies WHERE code = %s', (code,))
                if cursor.fetchone()['count'] > 0:
                    flash('Currency code already exists', 'error')
                    return redirect(url_for('admin_currencies'))
                
                # Insert new currency
                cursor.execute('''
                    INSERT INTO currencies (code, name, symbol, is_active) 
                    VALUES (%s, %s, %s, TRUE)
                ''', (code, name, symbol))
                
                # Get the new currency ID
                new_currency_id = cursor.lastrowid
                
                # Add exchange rates for the new currency with all active currencies
                cursor.execute('SELECT currency_id FROM currencies WHERE is_active = TRUE AND currency_id != %s', (new_currency_id,))
                active_currencies = cursor.fetchall()
                
                for curr in active_currencies:
                    other_id = curr['currency_id']
                    cursor.execute('''
                        INSERT INTO exchange_rates (from_currency_id, to_currency_id, rate)
                        VALUES (%s, %s, 1), (%s, %s, 1)
                    ''', (new_currency_id, other_id, other_id, new_currency_id))
                
                mysql.connection.commit()
                flash('Currency added successfully', 'success')
                return redirect(url_for('admin_currencies'))

            elif action == 'delete_currency':
                currency_id = request.form.get('currency_id')
                
                # Check if currency is in use
                cursor.execute('SELECT COUNT(*) as count FROM bookings WHERE currency_id = %s', (currency_id,))
                if cursor.fetchone()['count'] > 0:
                    flash('Cannot delete currency that is in use', 'error')
                    return redirect(url_for('admin_currencies'))
                
                # Delete related exchange rates first
                cursor.execute('DELETE FROM exchange_rates WHERE from_currency_id = %s OR to_currency_id = %s', 
                             (currency_id, currency_id))
                
                # Delete the currency
                cursor.execute('DELETE FROM currencies WHERE currency_id = %s', (currency_id,))
                
                mysql.connection.commit()
                flash('Currency deleted successfully', 'success')
                return redirect(url_for('admin_currencies'))
        
        # GET request - fetch currencies and rates
        cursor.execute('''
            SELECT 
                c.*,
                (SELECT COUNT(*) FROM bookings b WHERE b.currency_id = c.currency_id) as usage_count
            FROM currencies c
            ORDER BY c.code
        ''')
        currencies = cursor.fetchall()
        
        cursor.execute('''
            SELECT 
                er.*,
                fc.code as from_code,
                tc.code as to_code,
                fc.symbol as from_symbol,
                tc.symbol as to_symbol
            FROM exchange_rates er
            JOIN currencies fc ON er.from_currency_id = fc.currency_id
            JOIN currencies tc ON er.to_currency_id = tc.currency_id
            WHERE fc.is_active = TRUE AND tc.is_active = TRUE
            ORDER BY fc.code, tc.code
        ''')
        rates = cursor.fetchall()
        
        return render_template('admin/currencies.html', 
                             currencies=currencies,
                             rates=rates)
                             
    except Exception as e:
        app.logger.error(f"Admin currencies error: {str(e)}")
        flash('Error managing currencies', 'error')
        return redirect(url_for('admin_dashboard'))


# Manage gallery route handle gallery management for admin
@app.route('/admin/gallery')
@login_required
@admin_required
def manage_gallery():
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor) # Get database cursor
        
        # Pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = 10  # Number of images per page
        
        # Get filter parameters
        hotel_id = request.args.get('hotel_id')
        category = request.args.get('category')
        status = request.args.get('status')
        
        # Fetch hotels for dropdown
        cursor.execute("SELECT hotel_id, hotel_name FROM hotels ORDER BY hotel_name")
        hotels = cursor.fetchall()
        
        # Base count query
        count_query = '''
            SELECT COUNT(*) as total 
            FROM gallery g
            LEFT JOIN hotels h ON g.hotel_id = h.hotel_id
            WHERE 1=1
        '''
        
        # Base query
        query = '''
            SELECT g.*, h.hotel_name 
            FROM gallery g
            LEFT JOIN hotels h ON g.hotel_id = h.hotel_id
            WHERE 1=1
        '''
        params = []
        
        # Apply filters
        if hotel_id:
            query += ' AND g.hotel_id = %s' # Apply hotel filter
            count_query += ' AND g.hotel_id = %s' # Apply hotel filter
            params.append(hotel_id) # Add hotel ID to parameters

        # Apply category filter
        if category:
            query += ' AND g.category = %s' # Apply category filter
            count_query += ' AND g.category = %s' # Apply category filter
            params.append(category) # Add category to parameters
            
        if status:
            is_active = status == 'active'
            query += ' AND g.is_active = %s' # Apply status filter
            count_query += ' AND g.is_active = %s' # Apply status filter
            params.append(is_active) # Add status to parameters
        
        # Get total count
        cursor.execute(count_query, params)
        total_images = cursor.fetchone()['total']
        total_pages = (total_images + per_page - 1) // per_page
        
        # Add ordering and pagination
        query += ' ORDER BY g.display_order ASC LIMIT %s OFFSET %s' # Add ordering and pagination
        params.extend([per_page, (page - 1) * per_page]) # Add parameters
        
        # Execute main query
        cursor.execute(query, params) # Execute query
        gallery_images = cursor.fetchall() # Fetch all gallery images
        
        cursor.close()
        
        return render_template('admin/manage_gallery.html', 
                             gallery_images=gallery_images,
                             hotels=hotels,
                             page=page,
                             total_pages=total_pages,
                             total_images=total_images,
                             current_hotel=hotel_id,
                             current_category=category,
                             current_status=status)
                             
    except Exception as e:
        app.logger.error(f"Error fetching gallery: {str(e)}")
        flash('Error loading gallery images', 'danger')
        return redirect(url_for('admin_dashboard'))
    

# Add gallery image route handle image addition for admin
@app.route('/admin/gallery/add', methods=['POST'])
@login_required
@admin_required
def add_gallery_image():
    try:
        if 'image' not in request.files: # Check if image is uploaded
            flash('No image file uploaded', 'danger') # Show error message
            return redirect(url_for('manage_gallery')) # Redirect to gallery page
            
        file = request.files['image']
        if file.filename == '': # Check if no image is selected
            flash('No image selected', 'danger') # Show error message
            return redirect(url_for('manage_gallery')) # Redirect to gallery page
            
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename) # Secure filename
            filepath = os.path.join(app.config['GALLERY_IMAGES'], filename) # Save image to gallery folder
            file.save(filepath) # Save image
            
            cursor = mysql.connection.cursor() # Get database cursor
            cursor.execute("""
                INSERT INTO gallery (
                    title, 
                    image_url, 
                    category, 
                    description, 
                    hotel_id,
                    is_active
                )
                VALUES (%s, %s, %s, %s, %s, TRUE)
            """, (
                request.form['title'],
                f'/static/gallery_images/{filename}',
                request.form['category'],
                request.form['description'],
                request.form['hotel_id']
            ))
            mysql.connection.commit() # Commit changes
            cursor.close() # Close cursor
            
            flash('Image added successfully', 'success')
        else:
            flash('Invalid file type', 'danger')
            
        return redirect(url_for('manage_gallery'))
        
    except Exception as e:
        app.logger.error(f"Error adding gallery image: {str(e)}")
        flash('Error adding image', 'danger')
        return redirect(url_for('manage_gallery'))


# Get gallery image details route handle image details for admin
@app.route('/admin/gallery/<int:id>/details')
@login_required
@admin_required
def get_gallery_image_details(id):
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor) # Get database cursor
        
        # Fetch image details
        cursor.execute("""
            SELECT gallery_id, title, image_url, category, 
                   description, hotel_id, is_active
            FROM gallery 
            WHERE gallery_id = %s
        """, (id,))
        image = cursor.fetchone()
        cursor.close() # Close cursor
        
        if image:
            return jsonify({'success': True, 'image': image})
        return jsonify({'success': False, 'message': 'Image not found'})
        
    except Exception as e:
        app.logger.error(f"Error fetching image details: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/admin/gallery/<int:id>/edit', methods=['POST'])
@login_required
@admin_required
def edit_gallery_image(id):
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Check if image exists
        cursor.execute("""
            SELECT gallery_id, image_url, title, hotel_id 
            FROM gallery 
            WHERE gallery_id = %s
        """, (id,))
        existing_image = cursor.fetchone()
        
        if not existing_image:
            flash('Image not found', 'danger')
            return redirect(url_for('manage_gallery'))
        
        # Handle image upload if new image is provided
        image_url = existing_image['image_url']  # Keep existing image URL by default
        if 'image' in request.files and request.files['image'].filename:
            file = request.files['image']
            if allowed_file(file.filename):
                filename = secure_filename(f"{int(time.time())}_{file.filename}")
                filepath = os.path.join(app.config['GALLERY_IMAGES'], filename)
                file.save(filepath)
                image_url = f'/static/gallery_images/{filename}'
                
                # Delete old image file
                try:
                    old_path = os.path.join(
                        app.root_path, 
                        'static', 
                        existing_image['image_url'].lstrip('/static/')
                    )
                    if os.path.exists(old_path):
                        os.remove(old_path)
                except Exception as e:
                    app.logger.error(f"Error deleting old image: {str(e)}")

        # Update database
        cursor.execute("""
            UPDATE gallery 
            SET title = %s,
                image_url = %s,
                category = %s,
                description = %s,
                hotel_id = %s
            WHERE gallery_id = %s
        """, (
            request.form['title'],
            image_url,
            request.form['category'],
            request.form.get('description', ''),
            request.form['hotel_id'],
            id
        ))
        
        mysql.connection.commit()
        
        if cursor.rowcount > 0:
            flash('Image updated successfully', 'success')
        else:
            flash('No changes were made', 'info')
            
        return redirect(url_for('manage_gallery'))
        
    except Exception as e:
        app.logger.error(f"Error updating gallery image: {str(e)}")
        flash(f'Error updating image: {str(e)}', 'danger')
        return redirect(url_for('manage_gallery'))
    finally:
        if cursor:
            cursor.close()


# Delete gallery image route handle image deletion for admin
@app.route('/admin/gallery/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_gallery_image(id):
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor) # Get database cursor
        
        # First get the image details to delete the file
        cursor.execute("SELECT image_url FROM gallery WHERE gallery_id = %s", (id,))
        image = cursor.fetchone()
        
        if image:
            # Delete from database
            cursor.execute("DELETE FROM gallery WHERE gallery_id = %s", (id,))
            mysql.connection.commit()
            
            # Delete the physical file
            try:
                if image['image_url']:
                    file_path = os.path.join(
                        app.root_path,
                        'static',
                        image['image_url'].lstrip('/static/')
                    )
                    if os.path.exists(file_path):
                        os.remove(file_path)
            except Exception as e:
                app.logger.error(f"Error deleting file: {str(e)}")
                # Continue even if file deletion fails
            
            return jsonify({'success': True, 'message': 'Image deleted successfully'}) # Return success message
        
        return jsonify({'success': False, 'message': 'Image not found'}) # Return error message
        
    except Exception as e:
        app.logger.error(f"Error deleting gallery image: {str(e)}")
        return jsonify({'success': False, 'message': 'Error deleting image'})
    finally:
        cursor.close()


# destiantions route handle hotel list for user side
@app.route('/destinations')
def destinations():
    try:
        page = request.args.get('page', 1, type=int) # Get page number
        per_page = 9 # Number of hotels per page
        
        cursor = get_db() # Get database cursor
        
        # Fetch cities for dropdown
        cursor.execute('SELECT DISTINCT city FROM hotels WHERE status = "active"')
        cities = cursor.fetchall()
        
        # Fetch total number of hotels
        cursor.execute('SELECT COUNT(*) as total FROM hotels WHERE status = "active"')
        total_hotels = cursor.fetchone()['total']
        
        # Fetch hotels with amenities, starting price, and hero image
        cursor.execute('''
            SELECT 
                h.*,
                GROUP_CONCAT(DISTINCT ha.amenity_name) as amenities,
                GROUP_CONCAT(DISTINCT ha.icon_class) as amenity_icons,
                MIN(r.base_price) as starting_price
            FROM hotels h
            LEFT JOIN hotel_amenities ha ON h.hotel_id = ha.hotel_id
            LEFT JOIN rooms r ON h.hotel_id = r.hotel_id
            WHERE h.status = "active"
            GROUP BY h.hotel_id
            LIMIT %s OFFSET %s
        ''', (per_page, (page - 1) * per_page))
        
        hotels = cursor.fetchall()
        cursor.close()

        # If request is AJAX, return JSON instead of rendering template
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            hotels_list = []
            for hotel in hotels:
                hotel_dict = {}
                for key, value in dict(hotel).items():
                    if isinstance(value, (datetime, date)):
                        hotel_dict[key] = value.isoformat()
                    elif isinstance(value, timedelta):
                        total_seconds = int(value.total_seconds())
                        hotel_dict[key] = f"{total_seconds // 3600:02d}:{(total_seconds % 3600) // 60:02d}"
                    elif isinstance(value, Decimal):
                        hotel_dict[key] = float(value)
                    else:
                        hotel_dict[key] = value
                hotels_list.append(hotel_dict)

            return jsonify({
                'hotels': hotels_list,
                'hasMore': (page * per_page) < total_hotels
            })

        return render_template('destinations.html', 
                             hotels=hotels,
                             cities=cities,
                             total_hotels=total_hotels)

    except Exception as e:
        app.logger.error(f"Destinations page error: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': str(e)}), 500
        flash('Error loading destinations', 'error')
        return redirect(url_for('index'))



# hotels details route handle hotel details for user side after clicking on hotel
@app.route('/hotel/<int:hotel_id>')
def hotel_details(hotel_id):
    try:
        cursor = get_db()
        
        # Get hotel details with amenities, images, and hero image
        cursor.execute('''
            SELECT 
                h.*,
                GROUP_CONCAT(DISTINCT ha.amenity_name ORDER BY ha.amenity_id) as amenities,
                GROUP_CONCAT(DISTINCT ha.icon_class ORDER BY ha.amenity_id) as amenity_icons,
                MAX(rt.max_occupancy) as max_guests,
                (SELECT image_url FROM gallery 
                WHERE hotel_id = h.hotel_id 
                AND category = 'hero' 
                ORDER BY display_order LIMIT 1) as hero_image,
                (SELECT image_url FROM gallery 
                WHERE hotel_id = h.hotel_id 
                AND category = 'rooms' 
                ORDER BY display_order LIMIT 1) as featured_image
            FROM hotels h
            LEFT JOIN hotel_amenities ha ON h.hotel_id = ha.hotel_id
            LEFT JOIN rooms r ON h.hotel_id = r.hotel_id
            LEFT JOIN room_types rt ON r.room_type_id = rt.room_type_id
            WHERE h.hotel_id = %s
            GROUP BY h.hotel_id
        ''', (hotel_id,))
        hotel = cursor.fetchone()

        if not hotel:
            flash('Hotel not found', 'error')
            return redirect(url_for('destinations'))

        # Process amenities into a list of dicts
        amenities = []
        if hotel['amenities']:
            names = hotel['amenities'].split(',')
            icons = hotel['amenity_icons'].split(',')
            amenities = [{'name': n, 'icon': i} for n, i in zip(names, icons)]

        # Get available room types with features, amenities, and images
        cursor.execute('''
            SELECT DISTINCT 
                rt.type_name,
                rt.description,
                rt.max_occupancy,
                MIN(r.base_price) as starting_price,
                GROUP_CONCAT(DISTINCT rf.feature_name ORDER BY rf.feature_id) as features,
                GROUP_CONCAT(DISTINCT rf.icon_class ORDER BY rf.feature_id) as feature_icons,
                COUNT(DISTINCT r.room_id) as available_rooms,
                MIN(r.room_image) as room_image
            FROM room_types rt
            JOIN rooms r ON rt.room_type_id = r.room_type_id
            LEFT JOIN room_feature_mapping rfm ON r.room_id = rfm.room_id
            LEFT JOIN room_features rf ON rfm.feature_id = rf.feature_id
            WHERE r.hotel_id = %s 
            AND r.is_active = 1
            AND r.status = 'available'
            GROUP BY rt.room_type_id, rt.type_name, rt.description, rt.max_occupancy
        ''', (hotel_id,))
        room_types = cursor.fetchall()

        # Get available currencies
        cursor.execute('''
            SELECT 
                c.currency_id,
                c.code,
                c.symbol
            FROM currencies c
            WHERE c.is_active = 1
            ORDER BY c.code
        ''')
        currencies = cursor.fetchall()

        cursor.close()
        return render_template('hotel_details.html',
                             hotel=hotel,
                             amenities=amenities,
                             room_types=room_types,
                             currencies=currencies)

    except Exception as e:
        app.logger.error(f"Hotel details error: {str(e)}")
        flash('Error loading hotel details', 'error')
        return redirect(url_for('destinations'))


# about route handle about page for user side
@app.route('/about')
def about():
    return render_template('basic_pages/about.html')


# contact route handle contact page for user side
@app.route('/contact', methods=['GET', 'POST'])
def contact():
    form = ContactForm()
    if form.validate_on_submit():
        try:
            cursor = get_db()
            cursor.execute('''
                INSERT INTO contact_messages 
                (name, email, subject, message, created_at)
                VALUES (%s, %s, %s, %s, NOW())
            ''', (form.name.data, form.email.data, 
                 form.subject.data, form.message.data))
            mysql.connection.commit()
            
            # Get current year for email template
            current_year = datetime.now().year
            
            # Send confirmation email to user
            send_email(
                to=form.email.data,
                subject='Thank you for contacting World Hotels',
                template='email/contact_confirmation.html',
                name=form.name.data,
                message=form.message.data,
                subject_text=form.subject.data,
                now=datetime.now()
            )
            
            # Send notification email to admin
            send_email(
                to=app.config['MAIL_USERNAME'],
                subject=f'New Contact Form Submission: {form.subject.data}',
                template='email/admin_contact_notification.html',
                name=form.name.data,
                email=form.email.data,
                message=form.message.data,
                subject_text=form.subject.data,
                now=datetime.now()
            )
            
            flash('Your message has been sent successfully!', 'success')
            return redirect(url_for('contact'))
        except Exception as e:
            app.logger.error(f"Contact form error: {str(e)}")
            flash('An error occurred. Please try again.', 'error')
        finally:
            cursor.close()
    return render_template('basic_pages/contact.html', form=form)


# gallery route handle gallery page for user side
@app.route('/gallery')
def gallery():
    cursor = get_db()
    # Fetch gallery items from database
    query = """
        SELECT g.*, h.hotel_name 
        FROM gallery g
        LEFT JOIN hotels h ON g.hotel_id = h.hotel_id
        WHERE g.is_active = TRUE
        ORDER BY g.display_order, g.created_at DESC
    """
    cursor.execute(query)
    gallery_items = cursor.fetchall()
    
    # Get unique categories for filters
    cursor.execute("SELECT DISTINCT category FROM gallery WHERE is_active = TRUE")
    categories = cursor.fetchall()
    
    return render_template('basic_pages/gallery.html', 
                         gallery_items=gallery_items,
                         categories=categories)


# terms route handle terms page for user side
@app.route('/terms')
def terms():
    return render_template('basic_pages/terms.html')


# privacy route handle privacy page for user side
@app.route('/privacy')
def privacy():
    return render_template('basic_pages/privacy.html')


# cookie policy route handle cookie policy page for user side
@app.before_request
def check_cookie_preferences():
    if 'cookie_preferences_set' not in request.cookies:
        g.show_cookie_banner = True
    else:
        g.show_cookie_banner = False


# Inject cookie preferences into templates
@app.context_processor
def inject_cookie_preferences():
    return {
        'cookie_preferences': get_cookie_preferences(),
        'show_cookie_banner': getattr(g, 'show_cookie_banner', True)
    }


# Set cookie function
def set_cookie(response, key, value, days_expire=365):
    expire_date = datetime.now() + timedelta(days=days_expire)
    response.set_cookie(
        key,
        value=value,
        expires=expire_date,
        secure=True,  # Only sent over HTTPS
        httponly=True,  # Not accessible via JavaScript
        samesite='Lax'  # Protection against CSRF
    )
    return response


# Get cookie preferences
def get_cookie_preferences():
    return {
        'essential': request.cookies.get('essential_cookies', 'true'),
        'analytics': request.cookies.get('analytics_cookies', 'false'),
        'functional': request.cookies.get('functional_cookies', 'false'),
        'marketing': request.cookies.get('marketing_cookies', 'false')
    }


# cookie policy route handle cookie policy page for user side
@app.route('/cookie-policy')
def cookie_policy():
    cookie_preferences = get_cookie_preferences()
    return render_template(
        'basic_pages/cookie-policy.html',
        now=datetime.now(),
        cookie_preferences=cookie_preferences
    )

# Save cookie preferences route handle cookie preferences for user side
@app.route('/save-cookie-preferences', methods=['POST'])
def save_cookie_preferences():
    try:
        preferences = request.get_json() # Get cookie preferences from request
        response = make_response(jsonify({'status': 'success'}))
        
        # Set each cookie preference
        set_cookie(response, 'essential_cookies', str(preferences.get('essential', True)).lower())
        set_cookie(response, 'analytics_cookies', str(preferences.get('analytics', False)).lower())
        set_cookie(response, 'functional_cookies', str(preferences.get('functional', False)).lower())
        set_cookie(response, 'marketing_cookies', str(preferences.get('marketing', False)).lower())
        
        flash('Your cookie preferences have been saved.', 'success')
        return response
        
    except Exception as e:
        app.logger.error(f"Error saving cookie preferences: {str(e)}")
        flash('Failed to save preferences. Please try again.', 'error')
        return jsonify({'status': 'error', 'message': str(e)}), 500


# check cookie session
@app.route('/check-session')
def check_session():
    if 'user_id' in session: # Check if user is logged in
        return jsonify({ 
            'logged_in': True,
            'session_permanent': session.permanent,
            'session_lifetime': str(app.permanent_session_lifetime),
            'user_id': session['user_id'],
            'first_name': session['first_name']
        }) # Return true if user is logged in
    return jsonify({ # Return false if user is not logged in
        'logged_in': False
    })




# Run app
if __name__ == "__main__":
    app.run(debug=True)
